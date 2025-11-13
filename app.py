from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, abort
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from datetime import datetime
import io
import os
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from db_utils import get_db_connection, execute_query, fetch_one, fetch_all, close_connection
from auth_utils import login_required, role_required
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-change-in-production')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Note: Bootstrap is loaded via CDN in templates, no need for Flask-Bootstrap extension

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def log_activity(conn, user_id, activity_type, description, sample_id=None):
    """
    Log user activity to the database
    
    Args:
        conn: Database connection
        user_id: ID of the user performing the action
        activity_type: Type of activity performed
        description: Detailed description of the activity
        sample_id: Optional sample ID related to the activity
    """
    try:
        execute_query(conn,
            """INSERT INTO activity_logs (user_id, sample_id, activity_type, description, timestamp)
               VALUES (%s, %s, %s, %s, NOW())""",
            (user_id, sample_id, activity_type, description))
    except Exception as e:
        print(f"Error logging activity: {e}")

# ============================================================================
# CONTEXT PROCESSORS
# ============================================================================

@app.context_processor
def inject_user_data():
    """Inject user session data into all templates"""
    try:
        from auth_utils import get_session_context
        return get_session_context()
    except Exception as e:
        # Log the error for debugging
        print(f"Error in context processor: {e}")
        import traceback
        traceback.print_exc()
        # Return safe defaults to prevent server crashes when database is unavailable
        return {
            'is_authenticated': False,
            'current_user_id': None,
            'current_username': None,
            'current_user_role': None,
            'current_user_full_name': None,
            'is_admin': False,
            'is_personnel': False,
            'is_student': False,
            'current_user': None
        }

# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/')
def index():
    """Redirect to appropriate dashboard based on user role or login page"""
    if 'user_id' in session:
        role = session.get('role')
        if role == 'admin':
            return redirect(url_for('admin_dashboard'))
        elif role == 'personnel':
            return redirect(url_for('personnel_dashboard'))
        elif role == 'student':
            return redirect(url_for('student_dashboard'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Handle user login"""
    if request.method == 'POST':
        conn = None
        try:
            # Email-only login
            email = request.form.get('email')
            password = request.form.get('password')
            selected_role = request.form.get('organization', 'student')  # Get selected organization/role
            
            conn = get_db_connection()
            # Lookup by email only
            user = fetch_one(conn, 
                "SELECT * FROM users WHERE email = %s AND is_active = TRUE",
                (email,))
            
            if user and check_password_hash(user['password_hash'], password):
                # Check if user's actual role matches the selected role
                if user['role'] != selected_role:
                    close_connection(conn)
                    flash(f'Invalid role selection. Please select "{user["role"].title()}" for this account.', 'danger')
                    return render_template('login.html')
                
                # Update last login
                execute_query(conn,
                    "UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE user_id = %s",
                    (user['user_id'],))
                
                # Set session variables
                session['user_id'] = user['user_id']
                session['username'] = user['username']
                session['role'] = user['role']
                session['full_name'] = f"{user['first_name']} {user['last_name']}"
                
                close_connection(conn)
                flash('Login successful!', 'success')
                return redirect(url_for('index'))
            else:
                close_connection(conn)
                flash('Invalid credentials or inactive account', 'danger')
        except Exception as e:
            if conn:
                try:
                    close_connection(conn)
                except:
                    pass
            print(f"Error in login route: {e}")
            import traceback
            traceback.print_exc()
            flash('Database connection error. Please ensure XAMPP MySQL is running.', 'danger')
    
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    """Handle new user registration"""
    if request.method == 'POST':
        # Get form data
        first_name = request.form.get('first_name', '').strip()
        last_name = request.form.get('last_name', '').strip()
        email = request.form.get('email')
        password = request.form.get('password')
        student_id = request.form.get('student_id')
        
        # Validate first and last name
        if not first_name or not last_name:
            flash('First name and last name are required', 'danger')
            return render_template('signup.html')
        
        # Create username from email (before @)
        username = email.split('@')[0] if email else ''
        
        # Hash password
        password_hash = generate_password_hash(password)
        
        conn = get_db_connection()
        try:
            # Enforce uniqueness: prevent duplicate email or school_id
            existing_user = fetch_one(
                conn,
                """
                SELECT user_id, email, school_id
                FROM users
                WHERE email = %s OR school_id = %s
                """,
                (email, student_id),
            )

            if existing_user:
                # Build specific error message(s)
                duplicate_fields = []
                if existing_user.get('email') == email:
                    duplicate_fields.append('email')
                if existing_user.get('school_id') == student_id:
                    duplicate_fields.append('school ID')
                msg = ' and '.join(duplicate_fields).title() + ' already exists'
                close_connection(conn)
                flash(msg, 'danger')
                return render_template('signup.html')

            execute_query(conn,
                """INSERT INTO users (username, email, password_hash, first_name, last_name, 
                   role, school_id, is_active) 
                   VALUES (%s, %s, %s, %s, %s, 'student', %s, TRUE)""",
                (username, email, password_hash, first_name, last_name, student_id))
            close_connection(conn)
            flash('Registration successful! Please login.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            close_connection(conn)
            flash(f'Registration failed: Username or email already exists', 'danger')
    
    return render_template('signup.html')

@app.route('/logout')
def logout():
    """Handle user logout"""
    session.clear()
    flash('You have been logged out', 'info')
    return redirect(url_for('login'))

# ============================================================================
# STUDENT ROUTES
# ============================================================================

def ensure_user_photo_columns(conn):
    """
    Ensure the users table has columns for profile photo storage.
    Safe to call multiple times.
    """
    try:
        # MySQL/MariaDB support IF NOT EXISTS for ADD COLUMN
        execute_query(conn, """
            ALTER TABLE users
            ADD COLUMN IF NOT EXISTS profile_image LONGBLOB NULL,
            ADD COLUMN IF NOT EXISTS profile_image_mime VARCHAR(255) NULL,
            ADD COLUMN IF NOT EXISTS profile_image_name VARCHAR(255) NULL,
            ADD COLUMN IF NOT EXISTS profile_image_size INT NULL
        """)
    except Exception as _e:
        # If database version doesn't support IF NOT EXISTS per-column, attempt adding individually
        try:
            execute_query(conn, "ALTER TABLE users ADD COLUMN profile_image LONGBLOB NULL")
        except Exception:
            pass
        try:
            execute_query(conn, "ALTER TABLE users ADD COLUMN profile_image_mime VARCHAR(255) NULL")
        except Exception:
            pass
        try:
            execute_query(conn, "ALTER TABLE users ADD COLUMN profile_image_name VARCHAR(255) NULL")
        except Exception:
            pass
        try:
            execute_query(conn, "ALTER TABLE users ADD COLUMN profile_image_size INT NULL")
        except Exception:
            pass

def ensure_rock_location_columns(conn):
    """
    Ensure the rock_samples table has barangay and province columns, and drop deprecated fields.
    """
    try:
        execute_query(conn, """
            ALTER TABLE rock_samples
            ADD COLUMN IF NOT EXISTS barangay VARCHAR(255) NULL,
            ADD COLUMN IF NOT EXISTS province VARCHAR(255) NULL
        """)
    except Exception:
        try:
            execute_query(conn, "ALTER TABLE rock_samples ADD COLUMN barangay VARCHAR(255) NULL")
        except Exception:
            pass
        try:
            execute_query(conn, "ALTER TABLE rock_samples ADD COLUMN province VARCHAR(255) NULL")
        except Exception:
            pass
    # Drop deprecated outcrop_id column if it exists
    try:
        execute_query(conn, "ALTER TABLE rock_samples DROP COLUMN IF EXISTS outcrop_id")
    except Exception:
        # For MySQL versions without DROP COLUMN IF EXISTS
        try:
            execute_query(conn, "ALTER TABLE rock_samples DROP COLUMN outcrop_id")
        except Exception:
            pass

@app.route('/student/dashboard')
@login_required
@role_required('student')
def student_dashboard():
    """Student dashboard with statistics and recent submissions"""
    conn = None
    try:
        conn = get_db_connection()
        user_id = session['user_id']
        
        # Get student information
        student = fetch_one(conn,
            """SELECT first_name, last_name FROM users WHERE user_id = %s""",
            (user_id,))
        
        if not student:
            close_connection(conn)
            flash('Student information not found', 'danger')
            return redirect(url_for('logout'))
        
        # Get basic statistics - use IFNULL to handle NULL from SUM() on empty tables
        stats = fetch_one(conn,
            """SELECT 
                IFNULL(COUNT(*), 0) as total_submissions,
                IFNULL(SUM(CASE WHEN status = 'verified' THEN 1 ELSE 0 END), 0) as verified_count,
                IFNULL(SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END), 0) as pending_count,
                IFNULL(SUM(CASE WHEN status = 'rejected' THEN 1 ELSE 0 END), 0) as rejected_count
            FROM rock_samples WHERE user_id = %s""",
            (user_id,))
        
        # Get additional statistics
        additional_stats = fetch_one(conn,
            """SELECT 
                IFNULL(COUNT(DISTINCT rock_type), 0) as unique_rock_types,
                IFNULL(COUNT(DISTINCT location_name), 0) as unique_locations
            FROM rock_samples WHERE user_id = %s""",
            (user_id,))
        
        # Initialize stats safely
        if not stats:
            stats = {
                'total_submissions': 0,
                'verified_count': 0,
                'pending_count': 0,
                'rejected_count': 0
            }
        
        # Merge statistics safely
        if additional_stats:
            stats.update(additional_stats)
        else:
            stats.update({
                'unique_rock_types': 0,
                'unique_locations': 0
            })
        
        # Ensure all numeric values are integers (not None)
        for key in ['total_submissions', 'verified_count', 'pending_count', 'rejected_count', 
                    'unique_rock_types', 'unique_locations']:
            if key not in stats or stats[key] is None:
                stats[key] = 0
            else:
                stats[key] = int(stats[key])
        
        # Get recent submissions
        recent_rocks = fetch_all(conn,
            """SELECT * FROM rock_samples 
               WHERE user_id = %s 
               ORDER BY created_at DESC LIMIT 10""",
            (user_id,))
        
        close_connection(conn)
        return render_template('students/dashboard.html', 
                             stats=stats, 
                             recent_rocks=recent_rocks or [], 
                             student=student)
    except Exception as e:
        # Database error - close connection and show error
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Database error in student_dashboard: {e}")
        import traceback
        traceback.print_exc()
        flash('Database connection error. Please ensure XAMPP MySQL is running.', 'danger')
        return redirect(url_for('logout'))

@app.route('/student/add-rock', methods=['GET', 'POST'])
@login_required
@role_required('student')
def student_add_rock():
    """Add new rock sample submission"""
    if request.method == 'POST':
        user_id = session['user_id']
        rock_index = request.form.get('rock_index')
        rock_id = request.form.get('rock_id')
        rock_type = request.form.get('rock_type')
        description = request.form.get('description', '')
        formation = request.form.get('formation', '')
        location_name = request.form.get('location_name')
        barangay = request.form.get('barangay') or None
        province = request.form.get('province') or None
        
        # Handle optional latitude/longitude
        latitude_str = (request.form.get('latitude', '') or '').strip()
        longitude_str = (request.form.get('longitude', '') or '').strip()

        if not latitude_str or not longitude_str:
            flash('Latitude and longitude are required.', 'danger')
            return redirect(url_for('student_add_rock'))
        try:
            latitude = float(latitude_str)
            longitude = float(longitude_str)
        except ValueError:
            flash('Latitude and longitude must be numeric values.', 'danger')
            return redirect(url_for('student_add_rock'))

        if latitude < -90 or latitude > 90:
            flash('Latitude must be between -90 and 90 degrees.', 'danger')
            return redirect(url_for('student_add_rock'))
        if longitude < -180 or longitude > 180:
            flash('Longitude must be between -180 and 180 degrees.', 'danger')
            return redirect(url_for('student_add_rock'))
        
        rock_specimen = request.files.get('rock_specimen')
        outcrop_image = request.files.get('outcrop_image')
        
        conn = get_db_connection()
        try:
            ensure_rock_location_columns(conn)
            # Insert rock sample
            sample_id = execute_query(conn,
                """INSERT INTO rock_samples (user_id, rock_index, rock_id, rock_type, 
                   description, formation, location_name, barangay, province, latitude, longitude, 
                   status, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'pending', NOW())""",
                (user_id, rock_index, rock_id, rock_type, description, formation, 
                 location_name, barangay, province, latitude, longitude))
            
            # Insert rock specimen image (ensure only one image of this type)
            if rock_specimen and rock_specimen.filename:
                # Delete any existing rock_specimen images for this sample (safety check)
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'rock_specimen'""",
                    (sample_id,))
                rock_specimen_data = rock_specimen.read()
                if rock_specimen_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                           file_size, mime_type, created_at)
                           VALUES (%s, 'rock_specimen', %s, %s, %s, %s, NOW())""",
                        (sample_id, rock_specimen_data, rock_specimen.filename,
                             len(rock_specimen_data), rock_specimen.content_type or 'application/octet-stream'))
            
            # Insert outcrop image (ensure only one image of this type)
            if outcrop_image and outcrop_image.filename:
                # Delete any existing outcrop images for this sample (safety check)
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'outcrop'""",
                    (sample_id,))
                outcrop_image_data = outcrop_image.read()
                if outcrop_image_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                           file_size, mime_type, created_at)
                           VALUES (%s, 'outcrop', %s, %s, %s, %s, NOW())""",
                        (sample_id, outcrop_image_data, outcrop_image.filename,
                             len(outcrop_image_data), outcrop_image.content_type or 'application/octet-stream'))
            
            # Log activity
            log_activity(conn, user_id, 'submitted',
                        f'Submitted rock sample: {rock_id} - {rock_type}', sample_id)
            
            close_connection(conn)
            flash('Rock sample submitted successfully and is pending verification!', 'success')
            return redirect(url_for('student_view_rocks'))
        except Exception as e:
            close_connection(conn)
            flash(f'Error submitting rock sample: {str(e)}', 'danger')
    
    return render_template('students/add_rock.html')

@app.route('/student/edit-rock/<int:sample_id>', methods=['GET', 'POST'])
@login_required
@role_required('student')
def student_edit_rock(sample_id):
    """Edit rock sample belonging to the logged-in student."""
    conn = None
    user_id = session['user_id']
    
    try:
        conn = get_db_connection()
        ensure_rock_location_columns(conn)
        # First, verify the rock belongs to this student
        rock = fetch_one(conn,
            """SELECT * FROM rock_samples 
               WHERE sample_id = %s AND user_id = %s AND status IN ('pending', 'verified')""",
            (sample_id, user_id))
        active_nav = 'student_pending_verifications' if rock and rock.get('status') == 'pending' else 'student_view_rocks'
        
        if not rock:
            flash('Rock sample not found or cannot be edited.', 'danger')
            return redirect(url_for('student_pending_verifications'))
        
        if request.method == 'POST':
            rock_index = request.form.get('rock_index')
            rock_id = request.form.get('rock_id')
            rock_type = request.form.get('rock_type')
            description = request.form.get('description', '')
            formation = request.form.get('formation', '')
            location_name = request.form.get('location_name')
            barangay = request.form.get('barangay') or None
            province = request.form.get('province') or None
            
            # Handle optional latitude/longitude
            latitude_str = (request.form.get('latitude', '') or '').strip()
            longitude_str = (request.form.get('longitude', '') or '').strip()

            if not latitude_str or not longitude_str:
                flash('Latitude and longitude are required.', 'danger')
                return redirect(url_for('student_edit_rock', sample_id=sample_id))
            try:
                latitude = float(latitude_str)
                longitude = float(longitude_str)
            except ValueError:
                flash('Latitude and longitude must be numeric values.', 'danger')
                return redirect(url_for('student_edit_rock', sample_id=sample_id))

            if latitude < -90 or latitude > 90:
                flash('Latitude must be between -90 and 90 degrees.', 'danger')
                return redirect(url_for('student_edit_rock', sample_id=sample_id))
            if longitude < -180 or longitude > 180:
                flash('Longitude must be between -180 and 180 degrees.', 'danger')
                return redirect(url_for('student_edit_rock', sample_id=sample_id))
            
            rock_specimen = request.files.get('rock_specimen')
            outcrop_image = request.files.get('outcrop_image')
            
            try:
                current_status = rock['status']
                current_verified_by = rock.get('verified_by')
                # Update rock sample
                try:
                    execute_query(conn,
                        """UPDATE rock_samples 
                           SET rock_index = %s, rock_id = %s, rock_type = %s, 
                               description = %s, formation = %s, 
                               location_name = %s, barangay = %s, province = %s, latitude = %s, longitude = %s,
                               updated_at = NOW(),
                               status = %s,
                               verified_by = %s
                           WHERE sample_id = %s""",
                        (rock_index, rock_id, rock_type, description, formation, 
                         location_name, barangay, province, latitude, longitude,
                         current_status,
                         current_verified_by,
                         sample_id))
                except Exception:
                    # Fallback for schemas where verified_by cannot be NULL or updated
                    execute_query(conn,
                        """UPDATE rock_samples 
                           SET rock_index = %s, rock_id = %s, rock_type = %s, 
                               description = %s, formation = %s, 
                               location_name = %s, barangay = %s, province = %s, latitude = %s, longitude = %s,
                               updated_at = NOW(),
                               status = %s
                           WHERE sample_id = %s""",
                        (rock_index, rock_id, rock_type, description, formation, 
                         location_name, barangay, province, latitude, longitude,
                         current_status,
                         sample_id))

                # Update rock specimen image if provided (ensure only one image of this type)
                if rock_specimen and rock_specimen.filename:
                    # Delete existing rock specimen image before inserting new one
                    execute_query(conn,
                        """DELETE FROM images 
                           WHERE sample_id = %s AND image_type = 'rock_specimen'""",
                        (sample_id,))
                    rock_specimen_data = rock_specimen.read()
                    if rock_specimen_data:
                        # Insert new rock specimen image
                        execute_query(conn,
                            """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                               file_size, mime_type, created_at)
                               VALUES (%s, 'rock_specimen', %s, %s, %s, %s, NOW())""",
                            (sample_id, rock_specimen_data, rock_specimen.filename,
                                 len(rock_specimen_data), rock_specimen.content_type or 'application/octet-stream'))
                
                # Update outcrop image if provided (ensure only one image of this type)
                if outcrop_image and outcrop_image.filename:
                    # Delete existing outcrop image before inserting new one
                    execute_query(conn,
                        """DELETE FROM images 
                           WHERE sample_id = %s AND image_type = 'outcrop'""",
                        (sample_id,))
                    outcrop_image_data = outcrop_image.read()
                    if outcrop_image_data:
                        # Insert new outcrop image
                        execute_query(conn,
                            """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                               file_size, mime_type, created_at)
                               VALUES (%s, 'outcrop', %s, %s, %s, %s, NOW())""",
                            (sample_id, outcrop_image_data, outcrop_image.filename,
                                 len(outcrop_image_data), outcrop_image.content_type or 'application/octet-stream'))
                
                # Log activity
                log_activity(conn, user_id, 'updated', 
                            f'Updated rock sample: {rock_id} - {rock_type}', sample_id)
                
                flash('Rock sample updated successfully!', 'success')
                if current_status == 'pending':
                    return redirect(url_for('student_pending_verifications'))
                return redirect(url_for('student_view_rocks'))
            except Exception as e:
                flash(f'Error updating rock sample: {str(e)}', 'danger')
        
        # Get existing images for display
        images = fetch_all(conn,
            """SELECT image_id, image_type, file_name, file_size, mime_type, created_at
               FROM images WHERE sample_id = %s""",
            (sample_id,))
        
        return render_template(
            'students/edit_rock.html',
            rock=rock,
            images=images,
            active_nav=active_nav
        )
    finally:
        close_connection(conn)

@app.route('/student/delete-rock/<int:sample_id>', methods=['POST'])
@login_required
@role_required('student')
def student_delete_rock(sample_id):
    """Delete rock sample (students can only delete their own pending samples)"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    # First, verify the rock belongs to this student and is pending
    rock = fetch_one(conn,
        """SELECT * FROM rock_samples 
           WHERE sample_id = %s AND user_id = %s AND status = 'pending'""",
        (sample_id, user_id))
    
    if not rock:
        close_connection(conn)
        flash('Rock sample not found or cannot be deleted.', 'danger')
        return redirect(url_for('student_pending_verifications'))
    
    try:
        # Log the deletion activity before deleting other records
        execute_query(conn,
            """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
               VALUES (%s, %s, 'deleted', %s)""",
            (user_id, sample_id, f'Rock sample deleted: {rock["rock_id"]} - {rock["rock_type"]}'))
        
        # Delete associated images
        execute_query(conn,
            """DELETE FROM images WHERE sample_id = %s""",
            (sample_id,))
        
        # Delete the rock sample (activity logs will remain for history)
        execute_query(conn,
            """DELETE FROM rock_samples WHERE sample_id = %s""",
            (sample_id,))
        
        close_connection(conn)
        flash('Rock sample deleted successfully!', 'success')
        return redirect(url_for('student_pending_verifications'))
        
    except Exception as e:
        close_connection(conn)
        flash(f'Error deleting rock sample: {str(e)}', 'danger')
        return redirect(url_for('student_pending_verifications'))

@app.route('/student/view-rocks')
@login_required
@role_required('student')
def student_view_rocks():
    """View all verified rock samples from all students with search and filtering"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Get filter parameters from request
        search_query = request.args.get('search', '').strip()
        rock_type_filter = (request.args.get('rock_type', '') or '').strip()
        location_filter = (request.args.get('location', '') or '').strip()
        date_from = (request.args.get('date_from', '') or '').strip()
        date_to = (request.args.get('date_to', '') or '').strip()
        
        # Debug: Print filter parameters (remove in production)
        if rock_type_filter:
            print(f"DEBUG: rock_type_filter received: '{rock_type_filter}' (length: {len(rock_type_filter)})")
        
        # Build the base query - only verified rocks from all students
        base_query = """SELECT rs.*, 
            CONCAT(v.first_name, ' ', v.last_name) as verified_by_name,
            CONCAT(s.first_name, ' ', s.last_name) as submitted_by_name
            FROM rock_samples rs
            LEFT JOIN users v ON rs.verified_by = v.user_id
            LEFT JOIN users s ON rs.user_id = s.user_id
                    WHERE rs.status = 'verified'"""
        
        # Add WHERE conditions for filtering
        where_conditions = []
        params = []
        
        if search_query:
            where_conditions.append("(rs.rock_index LIKE %s OR rs.rock_id LIKE %s OR rs.rock_type LIKE %s OR rs.location_name LIKE %s OR rs.description LIKE %s)")
            search_param = f"%{search_query}%"
            params.extend([search_param, search_param, search_param, search_param, search_param])
        
        if rock_type_filter:
            where_conditions.append("rs.rock_type = %s")
            params.append(rock_type_filter)
        
        if location_filter:
            where_conditions.append("rs.location_name LIKE %s")
            params.append(f"%{location_filter}%")
        
        if date_from:
            where_conditions.append("DATE(rs.created_at) >= %s")
            params.append(date_from)
        
        if date_to:
            where_conditions.append("DATE(rs.created_at) <= %s")
            params.append(date_to)
        
        # Combine query with conditions
        if where_conditions:
            query = base_query + " AND " + " AND ".join(where_conditions) + " ORDER BY rs.created_at DESC"
        else:
            query = base_query + " ORDER BY rs.created_at DESC"
        
        # Execute query to get filtered rocks
        # Convert params list to tuple for MySQL connector compatibility, or None if empty
        query_params = tuple(params) if params else None
        
        # Debug: Print query details
        if rock_type_filter:
            print(f"DEBUG: Executing query with rock_type_filter='{rock_type_filter}'")
            print(f"DEBUG: Query: {query}")
            print(f"DEBUG: Params: {query_params}")
        
        try:
            rocks = fetch_all(conn, query, query_params)
            # Ensure rocks is a list, not None
            if rocks is None:
                rocks = []
            
            if rock_type_filter:
                print(f"DEBUG: Query executed successfully, found {len(rocks)} results")
                if len(rocks) > 0:
                    print(f"DEBUG: First rock: {rocks[0].get('rock_id')} - {rocks[0].get('rock_type')}")
        except Exception as e:
            print(f"DEBUG: Error executing query: {e}")
            print(f"DEBUG: Query: {query}")
            print(f"DEBUG: Params: {query_params}")
            import traceback
            traceback.print_exc()
            rocks = []
        
        # Get unique rock types and locations for filter dropdowns
        rock_types = fetch_all(conn, 
                    "SELECT DISTINCT rock_type FROM rock_samples WHERE status = 'verified' AND rock_type IS NOT NULL ORDER BY rock_type") or []
        
        locations = fetch_all(conn, 
                    "SELECT DISTINCT location_name FROM rock_samples WHERE status = 'verified' AND location_name IS NOT NULL ORDER BY location_name") or []
        
                # Query student's verified rocks directly from database
        my_verified = []
        if 'user_id' in session:
                    try:
                        user_id = int(session['user_id'])  # Ensure it's an integer
                        
                        # Build query for student's verified rocks with same filters
                        my_verified_base_query = """SELECT rs.*, 
                        CONCAT(v.first_name, ' ', v.last_name) as verified_by_name,
                        CONCAT(s.first_name, ' ', s.last_name) as submitted_by_name
                        FROM rock_samples rs
                        LEFT JOIN users v ON rs.verified_by = v.user_id
                        LEFT JOIN users s ON rs.user_id = s.user_id
                        WHERE rs.status = 'verified' AND rs.user_id = %s"""
                        
                        my_verified_where_conditions = []
                        my_verified_params = [user_id]
                        
                        if search_query:
                            my_verified_where_conditions.append("(rs.rock_index LIKE %s OR rs.rock_id LIKE %s OR rs.rock_type LIKE %s OR rs.location_name LIKE %s OR rs.description LIKE %s)")
                            search_param = f"%{search_query}%"
                            my_verified_params.extend([search_param, search_param, search_param, search_param, search_param])
                        
                        if rock_type_filter:
                            my_verified_where_conditions.append("rs.rock_type = %s")
                            my_verified_params.append(rock_type_filter)
                        
                        if location_filter:
                            my_verified_where_conditions.append("rs.location_name LIKE %s")
                            my_verified_params.append(f"%{location_filter}%")
                        
                        if date_from:
                            my_verified_where_conditions.append("DATE(rs.created_at) >= %s")
                            my_verified_params.append(date_from)
                        
                        if date_to:
                            my_verified_where_conditions.append("DATE(rs.created_at) <= %s")
                            my_verified_params.append(date_to)
                        
                        # Combine query with conditions
                        if my_verified_where_conditions:
                            my_verified_query = my_verified_base_query + " AND " + " AND ".join(my_verified_where_conditions) + " ORDER BY rs.created_at DESC"
                        else:
                            my_verified_query = my_verified_base_query + " ORDER BY rs.created_at DESC"
                        
                        # Convert params list to tuple for MySQL connector compatibility
                        my_verified_query_params = tuple(my_verified_params) if my_verified_params else None
                        my_verified = fetch_all(conn, my_verified_query, my_verified_query_params) or []
                    except Exception as e:
                        print(f"Error fetching student's verified rocks: {e}")
                        import traceback
                        traceback.print_exc()
                        my_verified = []
        
        close_connection(conn)
        
        # Debug: Verify rocks before rendering
        if rock_type_filter:
            print(f"DEBUG: Before rendering template - rocks count: {len(rocks) if rocks else 0}")
            print(f"DEBUG: rocks type: {type(rocks)}")
            if rocks and len(rocks) > 0:
                print(f"DEBUG: First rock sample_id: {rocks[0].get('sample_id')}")
                print(f"DEBUG: First rock rock_type: {rocks[0].get('rock_type')}")
        
        return render_template('students/view_rocks.html', rocks=rocks, 
                            search_query=search_query, rock_type_filter=rock_type_filter,
                            location_filter=location_filter, date_from=date_from, date_to=date_to,
                            rock_types=rock_types, locations=locations, my_verified=my_verified)
    except Exception as e:
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Error in student_view_rocks: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while loading rock samples.', 'danger')
        # Return empty data on error
        return render_template('students/view_rocks.html', rocks=[], 
                             search_query='', rock_type_filter='',
                             location_filter='', date_from='', date_to='',
                             rock_types=[], locations=[], my_verified=[])

def get_filtered_verified_rocks(conn, search_query='', rock_type_filter='', location_filter='', date_from='', date_to='', include_image_data=False):
    """
    Helper function to get filtered verified rocks based on search criteria.
    Returns the query, params tuple, and the rocks list.
    
    Args:
        conn: Database connection
        search_query: Search text
        rock_type_filter: Rock type filter
        location_filter: Location filter
        date_from: Start date filter
        date_to: End date filter
        include_image_data: If True, includes actual image data for Excel export
    
    Returns:
        List of rock dictionaries with image information
    """
    # Build the base query - only verified rocks from all students
    base_query = """SELECT rs.*, 
        CONCAT(v.first_name, ' ', v.last_name) as verified_by_name,
        CONCAT(s.first_name, ' ', s.last_name) as submitted_by_name
        FROM rock_samples rs
        LEFT JOIN users v ON rs.verified_by = v.user_id
        LEFT JOIN users s ON rs.user_id = s.user_id
        WHERE rs.status = 'verified'"""
    
    # Add WHERE conditions for filtering
    where_conditions = []
    params = []
    
    if search_query:
        where_conditions.append("(rs.rock_index LIKE %s OR rs.rock_id LIKE %s OR rs.rock_type LIKE %s OR rs.location_name LIKE %s OR rs.description LIKE %s)")
        search_param = f"%{search_query}%"
        params.extend([search_param, search_param, search_param, search_param, search_param])
    
    if rock_type_filter:
        where_conditions.append("rs.rock_type = %s")
        params.append(rock_type_filter)
    
    if location_filter:
        where_conditions.append("rs.location_name LIKE %s")
        params.append(f"%{location_filter}%")
    
    if date_from:
        where_conditions.append("DATE(rs.created_at) >= %s")
        params.append(date_from)
    
    if date_to:
        where_conditions.append("DATE(rs.created_at) <= %s")
        params.append(date_to)
    
    # Combine query with conditions
    if where_conditions:
        query = base_query + " AND " + " AND ".join(where_conditions) + " ORDER BY rs.created_at DESC"
    else:
        query = base_query + " ORDER BY rs.created_at DESC"
    
    # Execute query to get filtered rocks
    query_params = tuple(params) if params else None
    rocks = fetch_all(conn, query, query_params) or []
    
    # Get image information for all rocks
    if rocks:
        sample_ids = [rock.get('sample_id') for rock in rocks if rock.get('sample_id')]
        if sample_ids:
            # Initialize all rocks with image flags
            for rock in rocks:
                rock['has_rock_specimen'] = False
                rock['has_outcrop'] = False
                rock['rock_specimen_data'] = None
                rock['outcrop_data'] = None
                rock['rock_specimen_mime'] = None
                rock['outcrop_mime'] = None
            
            if include_image_data:
                # Fetch actual image data for Excel export
                placeholders = ','.join(['%s'] * len(sample_ids))
                images_query = f"""
                    SELECT sample_id, image_type, image_data, mime_type
                    FROM images
                    WHERE sample_id IN ({placeholders}) 
                    AND image_type IN ('rock_specimen', 'outcrop')
                    ORDER BY sample_id, image_type
                """
                image_results = fetch_all(conn, images_query, tuple(sample_ids))
                
                # Store image data in rocks - create a dict for faster lookup
                rock_dict = {rock.get('sample_id'): rock for rock in rocks}
                
                # Store image data in rocks
                for img in image_results:
                    sample_id = img.get('sample_id')
                    img_type = img.get('image_type')
                    if sample_id and img_type and sample_id in rock_dict:
                        rock = rock_dict[sample_id]
                        if img_type == 'rock_specimen':
                            rock['has_rock_specimen'] = True
                            rock['rock_specimen_data'] = img.get('image_data')
                            rock['rock_specimen_mime'] = img.get('mime_type')
                        elif img_type == 'outcrop':
                            rock['has_outcrop'] = True
                            rock['outcrop_data'] = img.get('image_data')
                            rock['outcrop_mime'] = img.get('mime_type')
            else:
                # Just check if images exist for CSV export
                placeholders = ','.join(['%s'] * len(sample_ids))
                images_query = f"""
                    SELECT sample_id, image_type
                    FROM images
                    WHERE sample_id IN ({placeholders}) 
                    AND image_type IN ('rock_specimen', 'outcrop')
                """
                image_results = fetch_all(conn, images_query, tuple(sample_ids))
                
                # Create a set for quick lookup
                image_set = set()
                for img in image_results:
                    sample_id = img.get('sample_id')
                    img_type = img.get('image_type')
                    if sample_id and img_type:
                        image_set.add((sample_id, img_type))
                
                # Set flags
                for rock in rocks:
                    sample_id = rock.get('sample_id')
                    if sample_id:
                        if (sample_id, 'rock_specimen') in image_set:
                            rock['has_rock_specimen'] = True
                        if (sample_id, 'outcrop') in image_set:
                            rock['has_outcrop'] = True
    
    return rocks

def get_filtered_personnel_rocks(conn, search_query='', rock_type_filter='', include_image_data=False):
    """
    Helper function to get filtered verified rocks for personnel based on search criteria.
    Excludes archived rocks.
    
    Args:
        conn: Database connection
        search_query: Search text
        rock_type_filter: Rock type filter (igneous, sedimentary, metamorphic)
        include_image_data: If True, includes actual image data for Excel export
    
    Returns:
        List of rock dictionaries with image information
    """
    # Build the base query - only verified rocks, excluding archived
    base_query = """SELECT rs.*, 
        CONCAT(u.first_name, ' ', u.last_name) as student_name,
        CONCAT(v.first_name, ' ', v.last_name) as verified_by_name
        FROM rock_samples rs
        JOIN users u ON rs.user_id = u.user_id
        LEFT JOIN users v ON rs.verified_by = v.user_id
        LEFT JOIN archives a ON rs.sample_id = a.sample_id
        WHERE rs.status = 'verified' AND a.sample_id IS NULL"""
    
    # Add WHERE conditions for filtering
    where_conditions = []
    params = []
    
    if search_query:
        where_conditions.append("""(rs.rock_id LIKE %s OR rs.rock_type LIKE %s 
                                   OR rs.location_name LIKE %s OR rs.description LIKE %s
                                   OR CONCAT(u.first_name, ' ', u.last_name) LIKE %s)""")
        search_param = f"%{search_query}%"
        params.extend([search_param] * 5)
    
    if rock_type_filter and rock_type_filter != '':
        # Map filter values to full rock type names
        rock_type_map = {
            'igneous': 'Igneous Rock',
            'sedimentary': 'Sedimentary Rock',
            'metamorphic': 'Metamorphic Rock'
        }
        if rock_type_filter.lower() in rock_type_map:
            where_conditions.append("rs.rock_type = %s")
            params.append(rock_type_map[rock_type_filter.lower()])
    
    # Combine query with conditions
    if where_conditions:
        query = base_query + " AND " + " AND ".join(where_conditions) + " ORDER BY rs.created_at DESC"
    else:
        query = base_query + " ORDER BY rs.created_at DESC"
    
    # Execute query to get filtered rocks
    query_params = tuple(params) if params else None
    rocks = fetch_all(conn, query, query_params) or []
    
    # Get image information for all rocks
    if rocks:
        sample_ids = [rock.get('sample_id') for rock in rocks if rock.get('sample_id')]
        if sample_ids:
            # Initialize all rocks with image flags
            for rock in rocks:
                rock['has_rock_specimen'] = False
                rock['has_outcrop'] = False
                rock['rock_specimen_data'] = None
                rock['outcrop_data'] = None
                rock['rock_specimen_mime'] = None
                rock['outcrop_mime'] = None
            
            if include_image_data:
                # Fetch actual image data for Excel export
                placeholders = ','.join(['%s'] * len(sample_ids))
                images_query = f"""
                    SELECT sample_id, image_type, image_data, mime_type
                    FROM images
                    WHERE sample_id IN ({placeholders}) 
                    AND image_type IN ('rock_specimen', 'outcrop')
                    ORDER BY sample_id, image_type
                """
                image_results = fetch_all(conn, images_query, tuple(sample_ids))
                
                # Store image data in rocks - create a dict for faster lookup
                rock_dict = {rock.get('sample_id'): rock for rock in rocks}
                
                # Store image data in rocks
                for img in image_results:
                    sample_id = img.get('sample_id')
                    img_type = img.get('image_type')
                    if sample_id and img_type and sample_id in rock_dict:
                        rock = rock_dict[sample_id]
                        if img_type == 'rock_specimen':
                            rock['has_rock_specimen'] = True
                            rock['rock_specimen_data'] = img.get('image_data')
                            rock['rock_specimen_mime'] = img.get('mime_type')
                        elif img_type == 'outcrop':
                            rock['has_outcrop'] = True
                            rock['outcrop_data'] = img.get('image_data')
                            rock['outcrop_mime'] = img.get('mime_type')
            else:
                # Just check if images exist for CSV export
                placeholders = ','.join(['%s'] * len(sample_ids))
                images_query = f"""
                    SELECT sample_id, image_type
                    FROM images
                    WHERE sample_id IN ({placeholders}) 
                    AND image_type IN ('rock_specimen', 'outcrop')
                """
                image_results = fetch_all(conn, images_query, tuple(sample_ids))
                
                # Create a set for quick lookup
                image_set = set()
                for img in image_results:
                    sample_id = img.get('sample_id')
                    img_type = img.get('image_type')
                    if sample_id and img_type:
                        image_set.add((sample_id, img_type))
                
                # Set flags
                for rock in rocks:
                    sample_id = rock.get('sample_id')
                    if sample_id:
                        if (sample_id, 'rock_specimen') in image_set:
                            rock['has_rock_specimen'] = True
                        if (sample_id, 'outcrop') in image_set:
                            rock['has_outcrop'] = True
    
    return rocks

def get_filtered_admin_rocks(conn, search_query='', rock_type_filter='', status_filter='', include_image_data=False):
    """
    Helper function to get filtered rocks for admin based on search criteria.
    Supports status filtering (verified, pending, rejected, or all).
    Excludes archived rocks.
    
    Args:
        conn: Database connection
        search_query: Search text
        rock_type_filter: Rock type filter
        status_filter: Status filter ('verified', 'pending', 'rejected', or '' for all)
        include_image_data: If True, includes actual image data for Excel export
    
    Returns:
        List of rock dictionaries with image information
    """
    # Build the base query - all rocks, excluding archived
    base_query = """SELECT rs.*, 
        CONCAT(u.first_name, ' ', u.last_name) as student_name,
        u.school_id as student_id,
        CONCAT(v.first_name, ' ', v.last_name) as verified_by_name
        FROM rock_samples rs
        JOIN users u ON rs.user_id = u.user_id
        LEFT JOIN users v ON rs.verified_by = v.user_id
        LEFT JOIN archives a ON rs.sample_id = a.sample_id
        WHERE a.sample_id IS NULL"""
    
    # Add WHERE conditions for filtering
    where_conditions = []
    params = []
    
    # Add status filter
    if status_filter and status_filter.lower() in ['verified', 'pending', 'rejected']:
        where_conditions.append("rs.status = %s")
        params.append(status_filter.lower())
    
    if search_query:
        where_conditions.append("""(rs.rock_id LIKE %s OR rs.rock_type LIKE %s 
                                   OR rs.location_name LIKE %s OR rs.description LIKE %s
                                   OR CONCAT(u.first_name, ' ', u.last_name) LIKE %s
                                   OR u.school_id LIKE %s)""")
        search_param = f"%{search_query}%"
        params.extend([search_param] * 6)
    
    if rock_type_filter and rock_type_filter != '':
        where_conditions.append("rs.rock_type = %s")
        params.append(rock_type_filter)
    
    # Combine query with conditions
    if where_conditions:
        query = base_query + " AND " + " AND ".join(where_conditions) + " ORDER BY rs.created_at DESC"
    else:
        query = base_query + " ORDER BY rs.created_at DESC"
    
    # Execute query to get filtered rocks
    query_params = tuple(params) if params else None
    rocks = fetch_all(conn, query, query_params) or []
    
    # Get image information for all rocks
    if rocks:
        sample_ids = [rock.get('sample_id') for rock in rocks if rock.get('sample_id')]
        if sample_ids:
            # Initialize all rocks with image flags
            for rock in rocks:
                rock['has_rock_specimen'] = False
                rock['has_outcrop'] = False
                rock['rock_specimen_data'] = None
                rock['outcrop_data'] = None
                rock['rock_specimen_mime'] = None
                rock['outcrop_mime'] = None
            
            if include_image_data:
                # Fetch actual image data for Excel export
                placeholders = ','.join(['%s'] * len(sample_ids))
                images_query = f"""
                    SELECT sample_id, image_type, image_data, mime_type
                    FROM images
                    WHERE sample_id IN ({placeholders}) 
                    AND image_type IN ('rock_specimen', 'outcrop')
                    ORDER BY sample_id, image_type
                """
                image_results = fetch_all(conn, images_query, tuple(sample_ids))
                
                # Store image data in rocks - create a dict for faster lookup
                rock_dict = {rock.get('sample_id'): rock for rock in rocks}
                
                # Store image data in rocks
                for img in image_results:
                    sample_id = img.get('sample_id')
                    img_type = img.get('image_type')
                    if sample_id and img_type and sample_id in rock_dict:
                        rock = rock_dict[sample_id]
                        if img_type == 'rock_specimen':
                            rock['has_rock_specimen'] = True
                            rock['rock_specimen_data'] = img.get('image_data')
                            rock['rock_specimen_mime'] = img.get('mime_type')
                        elif img_type == 'outcrop':
                            rock['has_outcrop'] = True
                            rock['outcrop_data'] = img.get('image_data')
                            rock['outcrop_mime'] = img.get('mime_type')
            else:
                # Just check if images exist for CSV export
                placeholders = ','.join(['%s'] * len(sample_ids))
                images_query = f"""
                    SELECT sample_id, image_type
                    FROM images
                    WHERE sample_id IN ({placeholders}) 
                    AND image_type IN ('rock_specimen', 'outcrop')
                """
                image_results = fetch_all(conn, images_query, tuple(sample_ids))
                
                # Create a set for quick lookup
                image_set = set()
                for img in image_results:
                    sample_id = img.get('sample_id')
                    img_type = img.get('image_type')
                    if sample_id and img_type:
                        image_set.add((sample_id, img_type))
                
                # Set flags
                for rock in rocks:
                    sample_id = rock.get('sample_id')
                    if sample_id:
                        if (sample_id, 'rock_specimen') in image_set:
                            rock['has_rock_specimen'] = True
                        if (sample_id, 'outcrop') in image_set:
                            rock['has_outcrop'] = True
    
    return rocks

@app.route('/student/export-rocks/csv')
@login_required
@role_required('student')
def student_export_rocks_csv():
    """Export verified rock samples to CSV format based on current filters"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Get filter parameters from request (same as view_rocks)
        search_query = request.args.get('search', '').strip()
        rock_type_filter = (request.args.get('rock_type', '') or '').strip()
        location_filter = (request.args.get('location', '') or '').strip()
        date_from = (request.args.get('date_from', '') or '').strip()
        date_to = (request.args.get('date_to', '') or '').strip()
        
        # Get filtered rocks (without image data for CSV)
        rocks = get_filtered_verified_rocks(conn, search_query, rock_type_filter, location_filter, date_from, date_to, include_image_data=False)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header row
        writer.writerow([
            'Sample Index', 'Rock ID', 'Rock Type', 'Description', 'Formation',
            'Location Name', 'Barangay', 'Province', 'Latitude', 'Longitude',
            'Rock Specimen Image', 'Outcrop Image',
            'Submitted By', 'Verified By', 'Status', 'Created At', 'Updated At'
        ])
        
        # Write data rows
        for rock in rocks:
            # Check if images exist and show "(image)" if they do
            rock_specimen_img = '(image)' if rock.get('has_rock_specimen') else ''
            outcrop_img = '(image)' if rock.get('has_outcrop') else ''
            
            writer.writerow([
                rock.get('rock_index', ''),
                rock.get('rock_id', ''),
                rock.get('rock_type', ''),
                rock.get('description', ''),
                rock.get('formation', ''),
                rock.get('location_name', ''),
                rock.get('barangay', ''),
                rock.get('province', ''),
                rock.get('latitude', ''),
                rock.get('longitude', ''),
                rock_specimen_img,
                outcrop_img,
                rock.get('submitted_by_name', ''),
                rock.get('verified_by_name', ''),
                rock.get('status', ''),
                rock.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('created_at') else '',
                rock.get('updated_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('updated_at') else ''
            ])
        
        # Prepare file for download
        output.seek(0)
        filename = f"verified_rocks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        close_connection(conn)
        
        # Create BytesIO object for CSV
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8'))
        mem.seek(0)
        
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Error exporting rocks to CSV: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while exporting data.', 'danger')
        return redirect(url_for('student_view_rocks'))

@app.route('/student/export-rocks/excel')
@login_required
@role_required('student')
def student_export_rocks_excel():
    """Export verified rock samples to Excel format based on current filters"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Get filter parameters from request (same as view_rocks)
        search_query = request.args.get('search', '').strip()
        rock_type_filter = (request.args.get('rock_type', '') or '').strip()
        location_filter = (request.args.get('location', '') or '').strip()
        date_from = (request.args.get('date_from', '') or '').strip()
        date_to = (request.args.get('date_to', '') or '').strip()
        
        # Get filtered rocks with image data for Excel export
        rocks = get_filtered_verified_rocks(conn, search_query, rock_type_filter, location_filter, date_from, date_to, include_image_data=True)
        
        # Create Excel workbook in memory
        wb = Workbook()
        ws = wb.active
        ws.title = "Verified Rock Samples"
        
        # Define header row (including image columns)
        headers = [
            'Sample Index', 'Rock ID', 'Rock Type', 'Description', 'Formation',
            'Location Name', 'Barangay', 'Province', 'Latitude', 'Longitude',
            'Rock Specimen Image', 'Outcrop Image',
            'Submitted By', 'Verified By', 'Status', 'Created At', 'Updated At'
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths for image columns (columns K and L) - wider for images
        ws.column_dimensions['K'].width = 25  # Rock Specimen Image
        ws.column_dimensions['L'].width = 25  # Outcrop Image
        
        # Set row height for image rows (higher to accommodate images)
        # Row height in Excel is in points (1 point = 1/72 inch), not pixels
        # For 150px images, we need approximately 112.5 points (150px * 0.75)
        image_row_height = 115  # points
        
        # Write data rows and embed images
        for row_num, rock in enumerate(rocks, 2):
            # Write text data
            ws.cell(row=row_num, column=1, value=rock.get('rock_index', ''))
            ws.cell(row=row_num, column=2, value=rock.get('rock_id', ''))
            ws.cell(row=row_num, column=3, value=rock.get('rock_type', ''))
            ws.cell(row=row_num, column=4, value=rock.get('description', ''))
            ws.cell(row=row_num, column=5, value=rock.get('formation', ''))
            ws.cell(row=row_num, column=6, value=rock.get('location_name', ''))
            ws.cell(row=row_num, column=7, value=rock.get('barangay', ''))
            ws.cell(row=row_num, column=8, value=rock.get('province', ''))
            ws.cell(row=row_num, column=9, value=rock.get('latitude', ''))
            ws.cell(row=row_num, column=10, value=rock.get('longitude', ''))
            
            # Set row height for images
            ws.row_dimensions[row_num].height = image_row_height
            
            # Column 11: Rock Specimen Image
            if rock.get('rock_specimen_data'):
                try:
                    # Create image from bytes
                    img_data = rock.get('rock_specimen_data')
                    # Ensure img_data is bytes
                    if not isinstance(img_data, bytes):
                        if isinstance(img_data, bytearray):
                            img_data = bytes(img_data)
                        else:
                            raise ValueError(f"Image data is not bytes, got {type(img_data)}")
                    img_bytes = io.BytesIO(img_data)
                    
                    # Resize image while maintaining aspect ratio using PIL
                    pil_img = PILImage.open(img_bytes)
                    max_size = 150
                    
                    # Store original dimensions
                    original_width = pil_img.width
                    original_height = pil_img.height
                    
                    # Calculate new size maintaining aspect ratio
                    if original_width > max_size or original_height > max_size:
                        # Use LANCZOS resampling for high quality
                        try:
                            pil_img.thumbnail((max_size, max_size), PILImage.Resampling.LANCZOS)
                        except (AttributeError, TypeError):
                            # Fallback for older PIL versions
                            pil_img.thumbnail((max_size, max_size), PILImage.LANCZOS)
                    
                    # Get final dimensions after thumbnail
                    final_width = pil_img.width
                    final_height = pil_img.height
                    
                    # Save resized image to BytesIO
                    resized_bytes = io.BytesIO()
                    # Determine best format - preserve PNG if possible, otherwise use JPEG
                    # OpenPyXL supports both JPEG and PNG
                    if pil_img.mode == 'RGBA':
                        # For RGBA, save as PNG to preserve transparency, or convert to RGB for JPEG
                        pil_img.save(resized_bytes, format='PNG')
                    elif pil_img.format and pil_img.format.upper() in ['PNG', 'JPEG', 'JPG']:
                        # Preserve original format if it's PNG or JPEG
                        pil_img.save(resized_bytes, format=pil_img.format, quality=85 if pil_img.format.upper() == 'JPEG' else None)
                    else:
                        # Convert other formats to JPEG
                        if pil_img.mode not in ['RGB', 'L']:
                            pil_img = pil_img.convert('RGB')
                        pil_img.save(resized_bytes, format='JPEG', quality=85)
                    resized_bytes.seek(0)
                    
                    # Create openpyxl image
                    img = OpenpyxlImage(resized_bytes)
                    # Set size - openpyxl uses pixels (approximately 96 DPI for Excel)
                    img.width = final_width
                    img.height = final_height
                    
                    # Anchor image to cell - center it in the cell
                    img.anchor = f'K{row_num}'
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error adding rock specimen image for sample {rock.get('sample_id')}: {e}")
                    import traceback
                    traceback.print_exc()
                    ws.cell(row=row_num, column=11, value='(image error)')
            else:
                ws.cell(row=row_num, column=11, value='')
            
            # Column 12: Outcrop Image
            if rock.get('outcrop_data'):
                try:
                    # Create image from bytes
                    img_data = rock.get('outcrop_data')
                    # Ensure img_data is bytes
                    if not isinstance(img_data, bytes):
                        if isinstance(img_data, bytearray):
                            img_data = bytes(img_data)
                        else:
                            raise ValueError(f"Image data is not bytes, got {type(img_data)}")
                    img_bytes = io.BytesIO(img_data)
                    
                    # Resize image while maintaining aspect ratio using PIL
                    pil_img = PILImage.open(img_bytes)
                    max_size = 150
                    
                    # Store original dimensions
                    original_width = pil_img.width
                    original_height = pil_img.height
                    
                    # Calculate new size maintaining aspect ratio
                    if original_width > max_size or original_height > max_size:
                        # Use LANCZOS resampling for high quality
                        try:
                            pil_img.thumbnail((max_size, max_size), PILImage.Resampling.LANCZOS)
                        except (AttributeError, TypeError):
                            # Fallback for older PIL versions
                            pil_img.thumbnail((max_size, max_size), PILImage.LANCZOS)
                    
                    # Get final dimensions after thumbnail
                    final_width = pil_img.width
                    final_height = pil_img.height
                    
                    # Save resized image to BytesIO
                    resized_bytes = io.BytesIO()
                    # Determine best format - preserve PNG if possible, otherwise use JPEG
                    # OpenPyXL supports both JPEG and PNG
                    if pil_img.mode == 'RGBA':
                        # For RGBA, save as PNG to preserve transparency, or convert to RGB for JPEG
                        pil_img.save(resized_bytes, format='PNG')
                    elif pil_img.format and pil_img.format.upper() in ['PNG', 'JPEG', 'JPG']:
                        # Preserve original format if it's PNG or JPEG
                        pil_img.save(resized_bytes, format=pil_img.format, quality=85 if pil_img.format.upper() == 'JPEG' else None)
                    else:
                        # Convert other formats to JPEG
                        if pil_img.mode not in ['RGB', 'L']:
                            pil_img = pil_img.convert('RGB')
                        pil_img.save(resized_bytes, format='JPEG', quality=85)
                    resized_bytes.seek(0)
                    
                    # Create openpyxl image
                    img = OpenpyxlImage(resized_bytes)
                    # Set size - openpyxl uses pixels (approximately 96 DPI for Excel)
                    img.width = final_width
                    img.height = final_height
                    
                    # Anchor image to cell - center it in the cell
                    img.anchor = f'L{row_num}'
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error adding outcrop image for sample {rock.get('sample_id')}: {e}")
                    import traceback
                    traceback.print_exc()
                    ws.cell(row=row_num, column=12, value='(image error)')
            else:
                ws.cell(row=row_num, column=12, value='')
            
            # Continue with remaining text columns
            ws.cell(row=row_num, column=13, value=rock.get('submitted_by_name', ''))
            ws.cell(row=row_num, column=14, value=rock.get('verified_by_name', ''))
            ws.cell(row=row_num, column=15, value=rock.get('status', ''))
            ws.cell(row=row_num, column=16, value=rock.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('created_at') else '')
            ws.cell(row=row_num, column=17, value=rock.get('updated_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('updated_at') else '')
        
        # Auto-adjust column widths for text columns (skip image columns K and L)
        for col_num in range(1, len(headers) + 1):
            if col_num not in [11, 12]:  # Skip image columns
                column_letter = ws.cell(row=1, column=col_num).column_letter
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=min(len(rocks) + 1, 100), min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                if adjusted_width > 0:
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"verified_rocks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        close_connection(conn)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Error exporting rocks to Excel: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while exporting data.', 'danger')
        return redirect(url_for('student_view_rocks'))

@app.route('/student/rock-detail/<int:rock_id>')
@login_required
@role_required('student')
def student_rock_detail(rock_id):
    """View details of rock sample - students can view their own rocks (any status) or verified rocks from others"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    # Get rock details - allow viewing:
    # 1. Student's own rocks (any status: pending, verified, rejected)
    # 2. Other students' verified rocks only
    rock = fetch_one(conn,
        """SELECT rs.*, 
           CONCAT(v.first_name, ' ', v.last_name) as verified_by_name,
           CONCAT(s.first_name, ' ', s.last_name) as submitted_by_name
           FROM rock_samples rs
           LEFT JOIN users v ON rs.verified_by = v.user_id
           LEFT JOIN users s ON rs.user_id = s.user_id
           WHERE rs.sample_id = %s 
           AND (rs.user_id = %s OR rs.status = 'verified')""",
        (rock_id, user_id))
    
    if not rock:
        flash('Rock sample not found or you do not have permission to view it', 'error')
        close_connection(conn)
        # Redirect based on whether it's their own rock or not
        # If it's a pending rock that doesn't belong to them, redirect to pending verifications
        # Otherwise redirect to view rocks
        return redirect(url_for('student_view_rocks'))
    
    # Get images for this rock sample
    images = fetch_all(conn,
        """SELECT image_id, image_type, file_name, file_size, mime_type, created_at
           FROM images WHERE sample_id = %s""",
        (rock_id,))
    
    close_connection(conn)
    return render_template('students/rock_detail.html', rock=rock, images=images)

@app.route('/student/pending-verifications')
@login_required
@role_required('student')
def student_pending_verifications():
    """View pending verifications for student"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    pending = fetch_all(conn,
        """SELECT *, DATEDIFF(NOW(), created_at) as days_pending 
           FROM rock_samples 
           WHERE user_id = %s AND status = 'pending'
           ORDER BY created_at DESC""",
        (user_id,))
    
    close_connection(conn)
    return render_template('students/pending_verifications.html', pending=pending)

@app.route('/student/archives')
@login_required
@role_required('student')
def student_archives():
    """View archived rock samples for student"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    archives = fetch_all(conn,
        """SELECT a.*, rs.rock_id, rs.rock_type, rs.location_name,
           CONCAT(u.first_name, ' ', u.last_name) as archived_by_name
           FROM archives a
           JOIN rock_samples rs ON a.sample_id = rs.sample_id
           JOIN users u ON a.archived_by = u.user_id
           WHERE rs.user_id = %s
           ORDER BY a.archived_at DESC""",
        (user_id,))
    
    close_connection(conn)
    return render_template('students/archives.html', archives=archives)

@app.route('/student/map')
@login_required
@role_required('student')
def student_map():
    """Interactive map showing verified rock sample locations from all students"""
    conn = get_db_connection()
    
    # Get all verified rock samples with coordinates from all students
    rocks = fetch_all(conn,
        """SELECT rs.sample_id, rs.rock_id, rs.rock_type, rs.formation, rs.location_name, 
           rs.latitude, rs.longitude, rs.status, rs.created_at,
           CONCAT(u.first_name, ' ', u.last_name) as student_name
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           WHERE rs.status = 'verified' AND rs.latitude IS NOT NULL AND rs.longitude IS NOT NULL
           ORDER BY rs.created_at DESC""")
    
    # Get city-level statistics for verified rocks only
    cities = fetch_all(conn,
        """SELECT location_name, COUNT(*) as specimen_count,
           AVG(latitude) as avg_lat, AVG(longitude) as avg_lng
           FROM rock_samples 
           WHERE status = 'verified' AND latitude IS NOT NULL AND longitude IS NOT NULL
           GROUP BY location_name""")
    
    close_connection(conn)
    return render_template('students/map.html', rocks=rocks, cities=cities)

@app.route('/student/logs')
@login_required
@role_required('student')
def student_logs():
    """View activity logs for student with filtering"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    # Get filter parameters from request
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # Build the base query
    base_query = """SELECT al.*, rs.rock_id, rs.rock_type, rs.status as sample_status
           FROM activity_logs al
           LEFT JOIN rock_samples rs ON al.sample_id = rs.sample_id
           WHERE al.user_id = %s"""
    
    # Add WHERE conditions for filtering
    where_conditions = [user_id]
    params = [user_id]
    
    if date_from:
        where_conditions.append("DATE(al.timestamp) >= %s")
        params.append(date_from)
    
    if date_to:
        where_conditions.append("DATE(al.timestamp) <= %s")
        params.append(date_to)
    
    # Combine query with conditions
    if len(where_conditions) > 1:
        query = base_query + " AND " + " AND ".join(where_conditions[1:]) + " ORDER BY al.timestamp DESC LIMIT 50"
    else:
        query = base_query + " ORDER BY al.timestamp DESC LIMIT 50"
    
    logs = fetch_all(conn, query, params)
    
    close_connection(conn)
    return render_template('students/logs.html', logs=logs, 
                         date_from=date_from, date_to=date_to)

@app.route('/student/settings', methods=['GET', 'POST'])
@login_required
@role_required('student')
def student_settings():
    """Student settings page"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    # Ensure columns exist for profile photos
    ensure_user_photo_columns(conn)
    
    if request.method == 'POST':
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        email = request.form.get('email')
        school_id = request.form.get('school_id')
        new_password = request.form.get('new_password')
        
        if new_password:
            password_hash = generate_password_hash(new_password)
            execute_query(conn,
                """UPDATE users SET first_name = %s, last_name = %s, email = %s, 
                   school_id = %s, password_hash = %s WHERE user_id = %s""",
                (first_name, last_name, email, school_id, password_hash, user_id))
        else:
            execute_query(conn,
                """UPDATE users SET first_name = %s, last_name = %s, email = %s, 
                   school_id = %s WHERE user_id = %s""",
                (first_name, last_name, email, school_id, user_id))
        
        flash('Settings updated successfully!', 'success')
        return redirect(url_for('student_settings'))
    
    user = fetch_one(conn, "SELECT * FROM users WHERE user_id = %s", (user_id,))
    close_connection(conn)
    return render_template('students/settings.html', user=user)

@app.route('/student/upload-photo', methods=['POST'])
@login_required
@role_required('student')
def student_upload_photo():
    """Upload or replace student profile photo"""
    conn = get_db_connection()
    try:
        ensure_user_photo_columns(conn)
        file = request.files.get('profile_photo')
        if not file or not file.filename:
            close_connection(conn)
            flash('Please select an image to upload', 'danger')
            return redirect(url_for('student_settings'))
        filename = secure_filename(file.filename)
        data = file.read()
        if not data:
            close_connection(conn)
            flash('Empty file uploaded', 'danger')
            return redirect(url_for('student_settings'))
        execute_query(conn, """
            UPDATE users
            SET profile_image = %s,
                profile_image_mime = %s,
                profile_image_name = %s,
                profile_image_size = %s,
                updated_at = NOW()
            WHERE user_id = %s
        """, (data, file.content_type or 'application/octet-stream', filename, len(data), session['user_id']))
        log_activity(conn, session['user_id'], 'profile_photo_updated', 'Updated profile photo')
        close_connection(conn)
        flash('Profile photo updated', 'success')
    except Exception as e:
        try:
            close_connection(conn)
        except Exception:
            pass
        flash(f'Error uploading photo: {str(e)}', 'danger')
    return redirect(url_for('student_settings'))

@app.route('/student/update-profile', methods=['POST'])
@login_required
@role_required('student')
def student_update_profile():
    """Update student profile"""
    try:
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        email = request.form.get('email')
        school_id = request.form.get('school_id')
        
        conn = get_db_connection()
        
        # Check if email is already taken by another user
        existing_user = fetch_one(conn,
            "SELECT user_id FROM users WHERE email = %s AND user_id != %s",
            (email, session['user_id']))
        
        if existing_user:
            flash('Email already taken by another user', 'error')
            close_connection(conn)
            return redirect(url_for('student_settings'))
        
        # Update user profile
        execute_query(conn,
            "UPDATE users SET first_name = %s, last_name = %s, email = %s, school_id = %s, updated_at = NOW() WHERE user_id = %s",
            (first_name, last_name, email, school_id, session['user_id']))
        
        # Update session
        session['full_name'] = f"{first_name} {last_name}"
        
        log_activity(conn, session['user_id'], 'profile_updated', 'Updated profile information')
        
        close_connection(conn)
        flash('Profile updated successfully', 'success')
        
    except Exception as e:
        flash(f'Error updating profile: {str(e)}', 'error')
        if 'conn' in locals():
            close_connection(conn)
    
    return redirect(url_for('student_settings'))

@app.route('/student/change-password', methods=['POST'])
@login_required
@role_required('student')
def student_change_password():
    """Change student password"""
    try:
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return redirect(url_for('student_settings'))
        if not new_password or len(new_password) < 8:
            flash('Password must be at least 8 characters long', 'error')
            return redirect(url_for('student_settings'))
        
        conn = get_db_connection()
        
        # Verify current password
        user = fetch_one(conn, "SELECT password_hash FROM users WHERE user_id = %s", (session['user_id'],))
        
        if not user or not check_password_hash(user['password_hash'], current_password):
            flash('Current password is incorrect', 'error')
            close_connection(conn)
            return redirect(url_for('student_settings'))
        
        # Update password
        new_hash = generate_password_hash(new_password)
        execute_query(conn,
            "UPDATE users SET password_hash = %s, updated_at = NOW() WHERE user_id = %s",
            (new_hash, session['user_id']))
        
        log_activity(conn, session['user_id'], 'password_changed', 'Changed password')
        
        close_connection(conn)
        flash('Password changed successfully', 'success')
        
    except Exception as e:
        flash(f'Error changing password: {str(e)}', 'error')
        if 'conn' in locals():
            close_connection(conn)
    
    return redirect(url_for('student_settings'))

@app.route('/student/update-notifications', methods=['POST'])
@login_required
@role_required('student')
def student_update_notifications():
    """Update student notification preferences"""
    try:
        # This is a placeholder for future notification functionality
        conn = get_db_connection()
        log_activity(conn, session['user_id'], 'notifications_updated', 'Updated notification preferences')
        close_connection(conn)
        
        flash('Notification preferences updated successfully', 'success')
        
    except Exception as e:
        flash(f'Error updating notifications: {str(e)}', 'error')
        if 'conn' in locals():
            close_connection(conn)
    
    return redirect(url_for('student_settings'))

# ============================================================================
# PERSONNEL ROUTES
# ============================================================================

@app.route('/personnel/settings')
@login_required
@role_required('personnel')
def personnel_settings():
    """Personnel settings page - view info and change password"""
    conn = get_db_connection()
    ensure_user_photo_columns(conn)
    user = fetch_one(conn, "SELECT * FROM users WHERE user_id = %s", (session['user_id'],))
    close_connection(conn)
    return render_template('personnel/settings.html', user=user)

@app.route('/personnel/upload-photo', methods=['POST'])
@login_required
@role_required('personnel')
def personnel_upload_photo():
    """Upload or replace personnel profile photo"""
    conn = get_db_connection()
    try:
        ensure_user_photo_columns(conn)
        file = request.files.get('profile_photo')
        if not file or not file.filename:
            close_connection(conn)
            flash('Please select an image to upload', 'danger')
            return redirect(url_for('personnel_settings'))
        filename = secure_filename(file.filename)
        data = file.read()
        if not data:
            close_connection(conn)
            flash('Empty file uploaded', 'danger')
            return redirect(url_for('personnel_settings'))
        execute_query(conn, """
            UPDATE users
            SET profile_image = %s,
                profile_image_mime = %s,
                profile_image_name = %s,
                profile_image_size = %s,
                updated_at = NOW()
            WHERE user_id = %s
        """, (data, file.content_type or 'application/octet-stream', filename, len(data), session['user_id']))
        log_activity(conn, session['user_id'], 'profile_photo_updated', 'Updated profile photo')
        close_connection(conn)
        flash('Profile photo updated', 'success')
    except Exception as e:
        try:
            close_connection(conn)
        except Exception:
            pass
        flash(f'Error uploading photo: {str(e)}', 'danger')
    return redirect(url_for('personnel_settings'))

@app.route('/personnel/change-password', methods=['POST'])
@login_required
@role_required('personnel')
def personnel_change_password():
    """Change personnel password"""
    try:
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return redirect(url_for('personnel_settings'))
        if not new_password or len(new_password) < 8:
            flash('Password must be at least 8 characters long', 'error')
            return redirect(url_for('personnel_settings'))
        
        conn = get_db_connection()
        user = fetch_one(conn, "SELECT password_hash FROM users WHERE user_id = %s", (session['user_id'],))
        if not user or not check_password_hash(user['password_hash'], current_password):
            flash('Current password is incorrect', 'error')
            close_connection(conn)
            return redirect(url_for('personnel_settings'))
        
        new_hash = generate_password_hash(new_password)
        execute_query(conn,
            "UPDATE users SET password_hash = %s, updated_at = NOW() WHERE user_id = %s",
            (new_hash, session['user_id']))
        
        log_activity(conn, session['user_id'], 'password_changed', 'Changed password')
        close_connection(conn)
        flash('Password changed successfully', 'success')
    except Exception as e:
        flash(f'Error changing password: {str(e)}', 'error')
        try:
            close_connection(conn)
        except Exception:
            pass
    return redirect(url_for('personnel_settings'))

@app.route('/personnel/dashboard')
@login_required
@role_required('personnel')
def personnel_dashboard():
    """Personnel dashboard with verification statistics"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    # Get comprehensive statistics
    stats = fetch_one(conn,
        """SELECT 
            (SELECT COUNT(*) FROM rock_samples) as total_rocks,
            (SELECT COUNT(*) FROM rock_samples WHERE status = 'pending') as pending_rocks,
            (SELECT COUNT(*) FROM rock_samples WHERE status = 'verified') as approved_rocks,
            (SELECT COUNT(*) FROM rock_samples WHERE status = 'rejected') as rejected_rocks,
            (SELECT COUNT(*) FROM rock_samples WHERE verified_by = %s) as verified_by_me
        """,
        (user_id,))
    
    # Get recent pending submissions for display
    recent_pending = fetch_all(conn,
        """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as student_name, u.school_id
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           WHERE rs.status = 'pending'
           ORDER BY rs.created_at ASC LIMIT 10""")
    
    close_connection(conn)
    return render_template('personnel/dashboard.html', 
                         total_rocks=stats['total_rocks'] or 0,
                         pending_rocks=stats['pending_rocks'] or 0,
                         approved_rocks=stats['approved_rocks'] or 0,
                         rejected_rocks=stats['rejected_rocks'] or 0,
                         recent_pending=recent_pending)

@app.route('/personnel/verification-panel')
@login_required
@role_required('personnel')
def personnel_verification_panel():
    """Verification panel for personnel to review rock samples"""
    conn = get_db_connection()
    
    pending_rocks = fetch_all(conn,
        """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as student_name,
           u.email as student_email, u.school_id
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           WHERE rs.status = 'pending'
           ORDER BY rs.created_at ASC""")
    
    close_connection(conn)
    return render_template('personnel/verification_panel.html', pending_rocks=pending_rocks)

@app.route('/personnel/verify-rock/<int:sample_id>', methods=['POST'])
@login_required
@role_required('personnel')
def personnel_verify_rock(sample_id):
    """Approve or reject a rock sample"""
    action = request.form.get('action')  # 'approve' or 'reject'
    remarks = request.form.get('remarks', '')
    user_id = session['user_id']
    
    conn = get_db_connection()
    
    if action == 'approve':
        execute_query(conn,
            """UPDATE rock_samples 
               SET status = 'verified', verified_by = %s, updated_at = CURRENT_TIMESTAMP
               WHERE sample_id = %s""",
            (user_id, sample_id))
        
        execute_query(conn,
            """INSERT INTO approval_logs (user_id, sample_id, action, remarks)
               VALUES (%s, %s, 'approved', %s)""",
            (user_id, sample_id, remarks))
        
        execute_query(conn,
            """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
               VALUES (%s, %s, 'approved', 'Rock sample approved')""",
            (user_id, sample_id))
        
        flash('Rock sample approved successfully!', 'success')
    
    elif action == 'reject':
        execute_query(conn,
            """UPDATE rock_samples 
               SET status = 'rejected', verified_by = %s, updated_at = CURRENT_TIMESTAMP
               WHERE sample_id = %s""",
            (user_id, sample_id))
        
        execute_query(conn,
            """INSERT INTO approval_logs (user_id, sample_id, action, remarks)
               VALUES (%s, %s, 'rejected', %s)""",
            (user_id, sample_id, remarks))
        
        execute_query(conn,
            """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
               VALUES (%s, %s, 'rejected', %s)""",
            (user_id, sample_id, f'Rock sample rejected: {remarks}'))
        
        flash('Rock sample rejected', 'warning')
    
    close_connection(conn)
    return redirect(url_for('personnel_verification_panel'))

@app.route('/personnel/rock-list')
@login_required
@role_required('personnel')
def personnel_rock_list():
    """View verified rock samples only with search and filtering"""
    conn = get_db_connection()
    
    # Get search and filter parameters
    search_query = request.args.get('search', '').strip()
    rock_type_filter = request.args.get('rock_type', '').strip()
    
    # Use helper function to get filtered rocks
    rocks = get_filtered_personnel_rocks(conn, search_query, rock_type_filter, include_image_data=False)
    
    close_connection(conn)
    return render_template('personnel/rock_list_view.html', rocks=rocks, 
                         search_query=search_query, rock_type_filter=rock_type_filter)

@app.route('/personnel/export-rocks/csv')
@login_required
@role_required('personnel')
def personnel_export_rocks_csv():
    """Export verified rock samples to CSV format based on current filters"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Get filter parameters from request (same as rock_list)
        search_query = request.args.get('search', '').strip()
        rock_type_filter = request.args.get('rock_type', '').strip()
        
        # Get filtered rocks (without image data for CSV)
        rocks = get_filtered_personnel_rocks(conn, search_query, rock_type_filter, include_image_data=False)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header row
        writer.writerow([
            'Sample Index', 'Rock ID', 'Rock Type', 'Description', 'Formation',
            'Location Name', 'Barangay', 'Province', 'Latitude', 'Longitude',
            'Rock Specimen Image', 'Outcrop Image',
            'Submitted By', 'Verified By', 'Status', 'Created At', 'Updated At'
        ])
        
        # Write data rows
        for rock in rocks:
            # Check if images exist and show "(image)" if they do
            rock_specimen_img = '(image)' if rock.get('has_rock_specimen') else ''
            outcrop_img = '(image)' if rock.get('has_outcrop') else ''
            
            writer.writerow([
                rock.get('rock_index', ''),
                rock.get('rock_id', ''),
                rock.get('rock_type', ''),
                rock.get('description', ''),
                rock.get('formation', ''),
                rock.get('location_name', ''),
                rock.get('barangay', ''),
                rock.get('province', ''),
                rock.get('latitude', ''),
                rock.get('longitude', ''),
                rock_specimen_img,
                outcrop_img,
                rock.get('student_name', ''),
                rock.get('verified_by_name', ''),
                rock.get('status', ''),
                rock.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('created_at') else '',
                rock.get('updated_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('updated_at') else ''
            ])
        
        # Prepare file for download
        output.seek(0)
        filename = f"verified_rocks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        close_connection(conn)
        
        # Create BytesIO object for CSV
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8'))
        mem.seek(0)
        
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Error exporting rocks to CSV: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while exporting data.', 'danger')
        return redirect(url_for('personnel_rock_list'))

@app.route('/personnel/export-rocks/excel')
@login_required
@role_required('personnel')
def personnel_export_rocks_excel():
    """Export verified rock samples to Excel format based on current filters"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Get filter parameters from request (same as rock_list)
        search_query = request.args.get('search', '').strip()
        rock_type_filter = request.args.get('rock_type', '').strip()
    
        # Get filtered rocks with image data for Excel export
        rocks = get_filtered_personnel_rocks(conn, search_query, rock_type_filter, include_image_data=True)
        
        # Create Excel workbook in memory
        wb = Workbook()
        ws = wb.active
        ws.title = "Verified Rock Samples"
        
        # Define header row (including image columns)
        headers = [
            'Sample Index', 'Rock ID', 'Rock Type', 'Description', 'Formation',
            'Location Name', 'Barangay', 'Province', 'Latitude', 'Longitude',
            'Rock Specimen Image', 'Outcrop Image',
            'Submitted By', 'Verified By', 'Status', 'Created At', 'Updated At'
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths for image columns (columns K and L) - wider for images
        ws.column_dimensions['K'].width = 25  # Rock Specimen Image
        ws.column_dimensions['L'].width = 25  # Outcrop Image
        
        # Set row height for image rows (higher to accommodate images)
        # Row height in Excel is in points (1 point = 1/72 inch), not pixels
        # For 150px images, we need approximately 112.5 points (150px * 0.75)
        image_row_height = 115  # points
        
        # Write data rows and embed images
        for row_num, rock in enumerate(rocks, 2):
            # Write text data
            ws.cell(row=row_num, column=1, value=rock.get('rock_index', ''))
            ws.cell(row=row_num, column=2, value=rock.get('rock_id', ''))
            ws.cell(row=row_num, column=3, value=rock.get('rock_type', ''))
            ws.cell(row=row_num, column=4, value=rock.get('description', ''))
            ws.cell(row=row_num, column=5, value=rock.get('formation', ''))
            ws.cell(row=row_num, column=6, value=rock.get('location_name', ''))
            ws.cell(row=row_num, column=7, value=rock.get('barangay', ''))
            ws.cell(row=row_num, column=8, value=rock.get('province', ''))
            ws.cell(row=row_num, column=9, value=rock.get('latitude', ''))
            ws.cell(row=row_num, column=10, value=rock.get('longitude', ''))
            
            # Set row height for images
            ws.row_dimensions[row_num].height = image_row_height
            
            # Column 11: Rock Specimen Image
            if rock.get('rock_specimen_data'):
                try:
                    # Create image from bytes
                    img_data = rock.get('rock_specimen_data')
                    # Ensure img_data is bytes
                    if not isinstance(img_data, bytes):
                        if isinstance(img_data, bytearray):
                            img_data = bytes(img_data)
                        else:
                            raise ValueError(f"Image data is not bytes, got {type(img_data)}")
                    img_bytes = io.BytesIO(img_data)
                    
                    # Resize image while maintaining aspect ratio using PIL
                    pil_img = PILImage.open(img_bytes)
                    max_size = 150
                    
                    # Store original dimensions
                    original_width = pil_img.width
                    original_height = pil_img.height
                    
                    # Calculate new size maintaining aspect ratio
                    if original_width > max_size or original_height > max_size:
                        # Use LANCZOS resampling for high quality
                        try:
                            pil_img.thumbnail((max_size, max_size), PILImage.Resampling.LANCZOS)
                        except (AttributeError, TypeError):
                            # Fallback for older PIL versions
                            pil_img.thumbnail((max_size, max_size), PILImage.LANCZOS)
                    
                    # Get final dimensions after thumbnail
                    final_width = pil_img.width
                    final_height = pil_img.height
                    
                    # Save resized image to BytesIO
                    resized_bytes = io.BytesIO()
                    # Determine best format - preserve PNG if possible, otherwise use JPEG
                    # OpenPyXL supports both JPEG and PNG
                    if pil_img.mode == 'RGBA':
                        # For RGBA, save as PNG to preserve transparency, or convert to RGB for JPEG
                        pil_img.save(resized_bytes, format='PNG')
                    elif pil_img.format and pil_img.format.upper() in ['PNG', 'JPEG', 'JPG']:
                        # Preserve original format if it's PNG or JPEG
                        pil_img.save(resized_bytes, format=pil_img.format, quality=85 if pil_img.format.upper() == 'JPEG' else None)
                    else:
                        # Convert other formats to JPEG
                        if pil_img.mode not in ['RGB', 'L']:
                            pil_img = pil_img.convert('RGB')
                        pil_img.save(resized_bytes, format='JPEG', quality=85)
                    resized_bytes.seek(0)
                    
                    # Create openpyxl image
                    img = OpenpyxlImage(resized_bytes)
                    # Set size - openpyxl uses pixels (approximately 96 DPI for Excel)
                    img.width = final_width
                    img.height = final_height
                    
                    # Anchor image to cell - center it in the cell
                    img.anchor = f'K{row_num}'
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error adding rock specimen image for sample {rock.get('sample_id')}: {e}")
                    import traceback
                    traceback.print_exc()
                    ws.cell(row=row_num, column=11, value='(image error)')
            else:
                ws.cell(row=row_num, column=11, value='')
            
            # Column 12: Outcrop Image
            if rock.get('outcrop_data'):
                try:
                    # Create image from bytes
                    img_data = rock.get('outcrop_data')
                    # Ensure img_data is bytes
                    if not isinstance(img_data, bytes):
                        if isinstance(img_data, bytearray):
                            img_data = bytes(img_data)
                        else:
                            raise ValueError(f"Image data is not bytes, got {type(img_data)}")
                    img_bytes = io.BytesIO(img_data)
                    
                    # Resize image while maintaining aspect ratio using PIL
                    pil_img = PILImage.open(img_bytes)
                    max_size = 150
                    
                    # Store original dimensions
                    original_width = pil_img.width
                    original_height = pil_img.height
                    
                    # Calculate new size maintaining aspect ratio
                    if original_width > max_size or original_height > max_size:
                        # Use LANCZOS resampling for high quality
                        try:
                            pil_img.thumbnail((max_size, max_size), PILImage.Resampling.LANCZOS)
                        except (AttributeError, TypeError):
                            # Fallback for older PIL versions
                            pil_img.thumbnail((max_size, max_size), PILImage.LANCZOS)
                    
                    # Get final dimensions after thumbnail
                    final_width = pil_img.width
                    final_height = pil_img.height
                    
                    # Save resized image to BytesIO
                    resized_bytes = io.BytesIO()
                    # Determine best format - preserve PNG if possible, otherwise use JPEG
                    # OpenPyXL supports both JPEG and PNG
                    if pil_img.mode == 'RGBA':
                        # For RGBA, save as PNG to preserve transparency, or convert to RGB for JPEG
                        pil_img.save(resized_bytes, format='PNG')
                    elif pil_img.format and pil_img.format.upper() in ['PNG', 'JPEG', 'JPG']:
                        # Preserve original format if it's PNG or JPEG
                        pil_img.save(resized_bytes, format=pil_img.format, quality=85 if pil_img.format.upper() == 'JPEG' else None)
                    else:
                        # Convert other formats to JPEG
                        if pil_img.mode not in ['RGB', 'L']:
                            pil_img = pil_img.convert('RGB')
                        pil_img.save(resized_bytes, format='JPEG', quality=85)
                    resized_bytes.seek(0)
                    
                    # Create openpyxl image
                    img = OpenpyxlImage(resized_bytes)
                    # Set size - openpyxl uses pixels (approximately 96 DPI for Excel)
                    img.width = final_width
                    img.height = final_height
                    
                    # Anchor image to cell - center it in the cell
                    img.anchor = f'L{row_num}'
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error adding outcrop image for sample {rock.get('sample_id')}: {e}")
                    import traceback
                    traceback.print_exc()
                    ws.cell(row=row_num, column=12, value='(image error)')
            else:
                ws.cell(row=row_num, column=12, value='')
            
            # Continue with remaining text columns
            ws.cell(row=row_num, column=13, value=rock.get('student_name', ''))
            ws.cell(row=row_num, column=14, value=rock.get('verified_by_name', ''))
            ws.cell(row=row_num, column=15, value=rock.get('status', ''))
            ws.cell(row=row_num, column=16, value=rock.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('created_at') else '')
            ws.cell(row=row_num, column=17, value=rock.get('updated_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('updated_at') else '')
        
        # Auto-adjust column widths for text columns (skip image columns K and L)
        for col_num in range(1, len(headers) + 1):
            if col_num not in [11, 12]:  # Skip image columns
                column_letter = ws.cell(row=1, column=col_num).column_letter
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=min(len(rocks) + 1, 100), min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                if adjusted_width > 0:
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"verified_rocks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
        close_connection(conn)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Error exporting rocks to Excel: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while exporting data.', 'danger')
        return redirect(url_for('personnel_rock_list'))

@app.route('/personnel/rock-detail/<int:sample_id>')
@login_required
@role_required('personnel')
def personnel_rock_detail(sample_id):
    """View detailed information about a rock sample"""
    conn = get_db_connection()
    
    rock = fetch_one(conn,
        """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as student_name,
           u.email as student_email, u.school_id,
           CONCAT(v.first_name, ' ', v.last_name) as verified_by_name
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           LEFT JOIN users v ON rs.verified_by = v.user_id
           WHERE rs.sample_id = %s""",
        (sample_id,))
    
    images = fetch_all(conn,
        """SELECT image_id, image_type, file_name, file_size, mime_type, created_at
           FROM images WHERE sample_id = %s""",
        (sample_id,))
    
    approval_history = fetch_all(conn,
        """SELECT al.*, CONCAT(u.first_name, ' ', u.last_name) as user_name
           FROM approval_logs al
           JOIN users u ON al.user_id = u.user_id
           WHERE al.sample_id = %s
           ORDER BY al.timestamp DESC""",
        (sample_id,))
    
    close_connection(conn)
    return render_template('personnel/rock_detail_view.html', rock=rock, 
                         images=images, approval_history=approval_history)

@app.route('/personnel/archive-rock/<int:sample_id>', methods=['POST'])
@login_required
@role_required('personnel')
def personnel_archive_rock(sample_id):
    """Archive a rock sample"""
    reason = request.form.get('reason', '')
    user_id = session['user_id']
    
    conn = get_db_connection()
    
    execute_query(conn,
        """INSERT INTO archives (sample_id, archived_by, archive_reason, status)
           VALUES (%s, %s, %s, 'archived')""",
        (sample_id, user_id, reason))
    
    execute_query(conn,
        """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
           VALUES (%s, %s, 'archived', %s)""",
        (user_id, sample_id, f'Rock sample archived: {reason}'))
    
    close_connection(conn)
    flash('Rock sample archived successfully!', 'success')
    return redirect(url_for('personnel_verification_panel'))

@app.route('/personnel/archives')
@login_required
@role_required('personnel')
def personnel_archives():
    """View archived rock samples - only show archives created by logged-in personnel"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    archives = fetch_all(conn,
        """SELECT a.*, rs.rock_id, rs.rock_type, rs.location_name,
           CONCAT(u.first_name, ' ', u.last_name) as archived_by_name,
           CONCAT(s.first_name, ' ', s.last_name) as student_name
           FROM archives a
           JOIN rock_samples rs ON a.sample_id = rs.sample_id
           JOIN users u ON a.archived_by = u.user_id
           JOIN users s ON rs.user_id = s.user_id
           WHERE a.archived_by = %s
           ORDER BY a.archived_at DESC""", (user_id,))
    
    close_connection(conn)
    return render_template('personnel/archives.html', archived_rocks=archives)

@app.route('/personnel/map')
@login_required
@role_required('personnel')
def personnel_map():
    """Interactive map showing all rock sample locations"""
    conn = get_db_connection()
    
    # Get all rock samples with coordinates
    rocks = fetch_all(conn,
        """SELECT rs.sample_id, rs.rock_id, rs.rock_type, rs.formation, rs.location_name, 
           rs.latitude, rs.longitude, rs.status, rs.created_at,
           CONCAT(u.first_name, ' ', u.last_name) as student_name
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           WHERE rs.latitude IS NOT NULL AND rs.longitude IS NOT NULL
           ORDER BY rs.created_at DESC""")
    
    # Get city-level statistics
    cities = fetch_all(conn,
        """SELECT location_name, COUNT(*) as specimen_count,
           AVG(latitude) as avg_lat, AVG(longitude) as avg_lng
           FROM rock_samples 
           WHERE latitude IS NOT NULL AND longitude IS NOT NULL
           GROUP BY location_name""")
    
    close_connection(conn)
    return render_template('personnel/map.html', rocks=rocks, cities=cities)

@app.route('/personnel/add-rock', methods=['GET', 'POST'])
@login_required
@role_required('personnel')
def personnel_add_rock():
    """Add new rock sample (personnel adds their own rock)"""
    if request.method == 'POST':
        user_id = session['user_id']  # Personnel's ID
        rock_index = request.form.get('rock_index')
        rock_id = request.form.get('rock_id')
        rock_type = request.form.get('rock_type')
        description = request.form.get('description', '')
        formation = request.form.get('formation', '')
        location_name = request.form.get('location_name')
        barangay = request.form.get('barangay') or None
        province = request.form.get('province') or None
        
        # Handle optional latitude/longitude
        latitude_str = (request.form.get('latitude', '') or '').strip()
        longitude_str = (request.form.get('longitude', '') or '').strip()

        if not latitude_str or not longitude_str:
            flash('Latitude and longitude are required.', 'danger')
            return redirect(url_for('personnel_add_rock'))
        try:
            latitude = float(latitude_str)
            longitude = float(longitude_str)
        except ValueError:
            flash('Latitude and longitude must be numeric values.', 'danger')
            return redirect(url_for('personnel_add_rock'))

        if latitude < -90 or latitude > 90:
            flash('Latitude must be between -90 and 90 degrees.', 'danger')
            return redirect(url_for('personnel_add_rock'))
        if longitude < -180 or longitude > 180:
            flash('Longitude must be between -180 and 180 degrees.', 'danger')
            return redirect(url_for('personnel_add_rock'))
        
        rock_specimen = request.files.get('rock_specimen')
        outcrop_image = request.files.get('outcrop_image')
        
        conn = get_db_connection()
        try:
            ensure_rock_location_columns(conn)
            # Insert rock sample - personnel rocks are auto-verified
            sample_id = execute_query(conn,
                """INSERT INTO rock_samples (user_id, rock_index, rock_id, rock_type, 
                   description, formation, location_name, barangay, province, latitude, longitude, 
                   status, verified_by, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'verified', %s, NOW())""",
                (user_id, rock_index, rock_id, rock_type, description, formation, 
                 location_name, barangay, province, latitude, longitude, user_id))
            
            # Insert rock specimen image (ensure only one image of this type)
            if rock_specimen and rock_specimen.filename:
                # Delete any existing rock_specimen images for this sample (safety check)
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'rock_specimen'""",
                    (sample_id,))
                rock_specimen_data = rock_specimen.read()
                if rock_specimen_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                           file_size, mime_type, created_at)
                           VALUES (%s, 'rock_specimen', %s, %s, %s, %s, NOW())""",
                        (sample_id, rock_specimen_data, rock_specimen.filename,
                             len(rock_specimen_data), rock_specimen.content_type or 'application/octet-stream'))
            
            # Insert outcrop image (ensure only one image of this type)
            if outcrop_image and outcrop_image.filename:
                # Delete any existing outcrop images for this sample (safety check)
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'outcrop'""",
                    (sample_id,))
                outcrop_image_data = outcrop_image.read()
                if outcrop_image_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                           file_size, mime_type, created_at)
                           VALUES (%s, 'outcrop', %s, %s, %s, %s, NOW())""",
                        (sample_id, outcrop_image_data, outcrop_image.filename,
                             len(outcrop_image_data), outcrop_image.content_type or 'application/octet-stream'))
            
            # Log activity
            log_activity(conn, session['user_id'], 'submitted', 
                        f'Personnel added rock sample: {rock_id} - {rock_type}', sample_id)
            
            close_connection(conn)
            flash('Rock sample added successfully and automatically verified!', 'success')
            return redirect(url_for('personnel_rock_list'))
        except Exception as e:
            close_connection(conn)
            flash(f'Error adding rock sample: {str(e)}', 'danger')
    
    return render_template('personnel/add_rock.html')

@app.route('/personnel/edit-rock/<int:sample_id>', methods=['GET', 'POST'])
@login_required
@role_required('personnel')
def personnel_edit_rock(sample_id):
    """Edit a rock sample"""
    conn = get_db_connection()
    ensure_rock_location_columns(conn)
    
    if request.method == 'POST':
        try:
            # Get form data
            rock_id = request.form.get('rock_id', '').strip()
            rock_type = request.form.get('rock_type', '').strip()
            description = request.form.get('description', '').strip()
            location_name = request.form.get('location_name', '').strip()
            barangay = request.form.get('barangay', '').strip() or None
            province = request.form.get('province', '').strip() or None
            latitude_str = (request.form.get('latitude', '') or '').strip()
            longitude_str = (request.form.get('longitude', '') or '').strip()
            formation = request.form.get('formation', '').strip()
            rock_index = request.form.get('rock_index', '').strip()

            if not latitude_str or not longitude_str:
                close_connection(conn)
                flash('Latitude and longitude are required.', 'error')
                return redirect(url_for('personnel_edit_rock', sample_id=sample_id))
            try:
                latitude = float(latitude_str)
                longitude = float(longitude_str)
            except ValueError:
                close_connection(conn)
                flash('Latitude and longitude must be numeric values.', 'error')
                return redirect(url_for('personnel_edit_rock', sample_id=sample_id))

            if latitude < -90 or latitude > 90:
                close_connection(conn)
                flash('Latitude must be between -90 and 90 degrees.', 'error')
                return redirect(url_for('personnel_edit_rock', sample_id=sample_id))
            if longitude < -180 or longitude > 180:
                close_connection(conn)
                flash('Longitude must be between -180 and 180 degrees.', 'error')
                return redirect(url_for('personnel_edit_rock', sample_id=sample_id))
            
            # Validate required fields
            if not rock_id or not rock_type or not location_name:
                flash('Rock ID, Rock Type, and Location are required!', 'error')
                return redirect(url_for('personnel_edit_rock', sample_id=sample_id))
            
            # Update the rock sample
            execute_query(conn,
                """UPDATE rock_samples SET 
                   rock_id = %s, rock_type = %s, description = %s,
                   location_name = %s, barangay = %s, province = %s, latitude = %s, longitude = %s,
                   formation = %s, rock_index = %s,
                   updated_at = NOW()
                   WHERE sample_id = %s""",
                (rock_id, rock_type, description, location_name, barangay, province, latitude, longitude,
                 formation, rock_index, sample_id))
            
            # Handle image removals
            remove_ids = request.form.getlist('remove_images')
            if remove_ids:
                for image_id in remove_ids:
                    if image_id.isdigit():
                        execute_query(conn,
                            "DELETE FROM images WHERE image_id = %s AND sample_id = %s",
                            (int(image_id), sample_id))

            # Handle new uploads - delete existing images of same type first to ensure only one
            rock_specimen_file = request.files.get('rock_specimen')
            if rock_specimen_file and rock_specimen_file.filename:
                # Delete existing rock_specimen image before inserting new one
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'rock_specimen'""",
                    (sample_id,))
                specimen_data = rock_specimen_file.read()
                if specimen_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name,
                           file_size, mime_type, created_at)
                           VALUES (%s, 'rock_specimen', %s, %s, %s, %s, NOW())""",
                        (sample_id, specimen_data, rock_specimen_file.filename,
                         len(specimen_data), rock_specimen_file.content_type or 'application/octet-stream'))

            outcrop_file = request.files.get('outcrop_image')
            if outcrop_file and outcrop_file.filename:
                # Delete existing outcrop image before inserting new one
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'outcrop'""",
                    (sample_id,))
                outcrop_data = outcrop_file.read()
                if outcrop_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name,
                           file_size, mime_type, created_at)
                           VALUES (%s, 'outcrop', %s, %s, %s, %s, NOW())""",
                        (sample_id, outcrop_data, outcrop_file.filename,
                         len(outcrop_data), outcrop_file.content_type or 'application/octet-stream'))
            
            # Log the edit activity
            user_id = session['user_id']
            execute_query(conn,
                """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
                   VALUES (%s, %s, 'edited', %s)""",
                (user_id, sample_id, f'Rock sample edited by personnel: {rock_id}'))
            
            close_connection(conn)
            flash('Rock sample updated successfully!', 'success')
            return redirect(url_for('personnel_rock_detail', sample_id=sample_id))
            
        except Exception as e:
            print(f"Error updating rock sample {sample_id}: {str(e)}")
            flash('Error updating rock sample!', 'error')
            close_connection(conn)
            return redirect(url_for('personnel_edit_rock', sample_id=sample_id))
    
    else:
        # GET request - show edit form
        rock = fetch_one(conn,
            """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as student_name,
               u.email as student_email, u.school_id
               FROM rock_samples rs
               JOIN users u ON rs.user_id = u.user_id
               WHERE rs.sample_id = %s""",
            (sample_id,))
        
        if not rock:
            close_connection(conn)
            flash('Rock sample not found!', 'error')
            return redirect(url_for('personnel_rock_list'))
        
        # Fetch associated images for preview in edit screen
        images = fetch_all(conn,
            """SELECT image_id, image_type, file_name, file_size, mime_type, created_at
               FROM images
               WHERE sample_id = %s
               ORDER BY created_at DESC""",
            (sample_id,))
        
        close_connection(conn)
        return render_template('personnel/edit_rock.html', rock=rock, images=images)

@app.route('/personnel/activity-logs')
@login_required
@role_required('personnel')
def personnel_activity_logs():
    """View activity logs with filtering - only show activities performed by logged-in personnel"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    # Get filter parameters from request
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    
    # Build the base query - filter by logged-in personnel user
    base_query = """SELECT al.*, CONCAT(u.first_name, ' ', u.last_name) as user_name,
           u.username, u.role as user_role, rs.rock_id, rs.rock_type
           FROM activity_logs al
           JOIN users u ON al.user_id = u.user_id
           LEFT JOIN rock_samples rs ON al.sample_id = rs.sample_id
           WHERE al.user_id = %s"""
    
    # Add WHERE conditions for filtering
    where_conditions = ["al.user_id = %s"]
    params = [user_id]
    
    if date_from:
        where_conditions.append("DATE(al.timestamp) >= %s")
        params.append(date_from)
    
    if date_to:
        where_conditions.append("DATE(al.timestamp) <= %s")
        params.append(date_to)
    
    # Combine query with conditions
    query = base_query + " AND " + " AND ".join(where_conditions[1:]) + " ORDER BY al.timestamp DESC LIMIT 100" if len(where_conditions) > 1 else base_query + " ORDER BY al.timestamp DESC LIMIT 100"
    
    logs = fetch_all(conn, query, params)
    
    close_connection(conn)
    return render_template('personnel/activity_logs.html', logs=logs, 
                         date_from=date_from, date_to=date_to)

# Removed export-to-CSV endpoint for personnel

# ============================================================================
# ADMIN ROUTES
# ============================================================================

@app.route('/admin/dashboard')
@login_required
@role_required('admin')
def admin_dashboard():
    """Admin dashboard with system statistics"""
    conn = get_db_connection()
    
    # Get comprehensive statistics
    stats = fetch_one(conn,
        """SELECT 
            (SELECT COUNT(*) FROM users WHERE is_active = TRUE) as total_users,
            (SELECT COUNT(*) FROM users WHERE role = 'student' AND is_active = TRUE) as total_students,
            (SELECT COUNT(*) FROM users WHERE role = 'personnel' AND is_active = TRUE) as total_personnel,
            (SELECT COUNT(*) FROM rock_samples) as total_rocks,
            (SELECT COUNT(*) FROM rock_samples WHERE status = 'verified') as verified_rocks,
            (SELECT COUNT(*) FROM rock_samples WHERE status = 'pending') as pending_rocks,
            (SELECT COUNT(*) FROM rock_samples WHERE status = 'rejected') as rejected_rocks,
            (SELECT COUNT(*) FROM archives) as archived_rocks""")
    
    # Get recent activity
    recent_activity = fetch_all(conn,
        """SELECT al.*, CONCAT(u.first_name, ' ', u.last_name) as user_name,
           u.role, rs.rock_id
           FROM activity_logs al
           JOIN users u ON al.user_id = u.user_id
           LEFT JOIN rock_samples rs ON al.sample_id = rs.sample_id
           ORDER BY al.timestamp DESC LIMIT 10""")
    
    # Get all rock samples for the dashboard table
    all_rocks = fetch_all(conn,
        """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as submitted_by_name,
           u.school_id as student_id
           FROM rock_samples rs
           LEFT JOIN users u ON rs.user_id = u.user_id
           ORDER BY rs.created_at DESC LIMIT 20""")
    
    close_connection(conn)
    return render_template('admin/dashboard.html', stats=stats, recent_activity=recent_activity, all_rocks=all_rocks)

@app.route('/admin/manage-users')
@login_required
@role_required('admin')
def admin_manage_users():
    """Manage all users in the system"""
    conn = get_db_connection()
    role_filter = request.args.get('role', '').strip()
    
    if role_filter in ('admin', 'personnel', 'student'):
        users = fetch_all(conn,
            """SELECT user_id, username, email, first_name, last_name, role, 
               school_id, is_active, created_at, last_login
               FROM users
               WHERE role = %s
               ORDER BY created_at DESC""",
            (role_filter,))
    else:
        users = fetch_all(conn,
            """SELECT user_id, username, email, first_name, last_name, role, 
               school_id, is_active, created_at, last_login
               FROM users
               ORDER BY created_at DESC""")
    
    close_connection(conn)
    return render_template('admin/manage_users.html', users=users, role_filter=role_filter)

@app.route('/admin/toggle-user/<int:user_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_toggle_user(user_id):
    """Activate or deactivate a user"""
    conn = get_db_connection()
    
    user = fetch_one(conn, "SELECT is_active FROM users WHERE user_id = %s", (user_id,))
    new_status = not user['is_active']
    
    execute_query(conn,
        "UPDATE users SET is_active = %s WHERE user_id = %s",
        (new_status, user_id))
    
    close_connection(conn)
    flash(f'User {"activated" if new_status else "deactivated"} successfully!', 'success')
    return redirect(url_for('admin_manage_users'))

@app.route('/admin/delete-user/<int:user_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_delete_user(user_id):
    """Delete a user from the system"""
    from flask import jsonify
    if user_id == session['user_id']:
        # Never allow self-delete
        return jsonify({ 'success': False, 'message': 'Cannot delete your own account.' }), 400
    
    try:
        conn = get_db_connection()
        # Optionally ensure user exists
        existing = fetch_one(conn, "SELECT user_id FROM users WHERE user_id = %s", (user_id,))
        if not existing:
            close_connection(conn)
            return jsonify({ 'success': False, 'message': 'User not found.' }), 404
        
        # To preserve rock samples, do NOT hard-delete users. Soft-disable instead.
        execute_query(conn,
            "UPDATE users SET is_active = 0, updated_at = NOW() WHERE user_id = %s",
            (user_id,))
        close_connection(conn)
        return jsonify({ 'success': True, 'message': 'User deactivated to preserve submitted rock records.' })
    except Exception as e:
        try:
            close_connection(conn)
        except Exception:
            pass
        return jsonify({ 'success': False, 'message': f'Error deleting user: {str(e)}' }), 500

@app.route('/admin/rock-list')
@login_required
@role_required('admin')
def admin_rock_list():
    """View all rock samples with search and filtering"""
    conn = get_db_connection()
    
    # Get search and filter parameters
    search_query = request.args.get('search', '').strip()
    rock_type_filter = request.args.get('rock_type', '').strip()
    
    # Debug logging (remove in production)
    print(f"Search query: '{search_query}', Rock type filter: '{rock_type_filter}'")
    
    # Build the base query
    base_query = """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as student_name,
           u.school_id as student_id,
           CONCAT(v.first_name, ' ', v.last_name) as verified_by_name
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           LEFT JOIN users v ON rs.verified_by = v.user_id
           LEFT JOIN archives a ON rs.sample_id = a.sample_id
           WHERE a.sample_id IS NULL"""
    
    # Build WHERE conditions
    where_conditions = []
    params = []
    
    # Add search functionality
    if search_query:
        where_conditions.append("""(rs.rock_id LIKE %s OR rs.rock_type LIKE %s 
                                   OR rs.location_name LIKE %s OR rs.description LIKE %s
                                   OR CONCAT(u.first_name, ' ', u.last_name) LIKE %s
                                   OR u.school_id LIKE %s)""")
        search_param = f"%{search_query}%"
        params.extend([search_param] * 6)  # 6 placeholders for the search
    
    # Add rock type filter
    if rock_type_filter and rock_type_filter != '':
        where_conditions.append("rs.rock_type = %s")
        params.append(rock_type_filter)
    
    # Combine query
    if where_conditions:
        query = base_query + " AND " + " AND ".join(where_conditions)
    else:
        query = base_query
    
    query += " ORDER BY rs.created_at DESC"
    
    # Debug logging (remove in production)
    print(f"Final query: {query}")
    print(f"Parameters: {params}")
    
    # Execute query
    rocks = fetch_all(conn, query, params)
    
    close_connection(conn)
    return render_template('admin/rock_list_view.html', rocks=rocks, 
                          search_query=search_query, rock_type_filter=rock_type_filter)

@app.route('/admin/export-rocks/csv')
@login_required
@role_required('admin')
def admin_export_rocks_csv():
    """Export rock samples to CSV format based on current filters and status"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Get filter parameters from request
        search_query = request.args.get('search', '').strip()
        rock_type_filter = request.args.get('rock_type', '').strip()
        status_filter = request.args.get('status', '').strip()  # verified, pending, rejected, or empty for all
        
        # Get filtered rocks (without image data for CSV)
        rocks = get_filtered_admin_rocks(conn, search_query, rock_type_filter, status_filter, include_image_data=False)
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header row
        writer.writerow([
            'Sample Index', 'Rock ID', 'Rock Type', 'Description', 'Formation',
            'Location Name', 'Barangay', 'Province', 'Latitude', 'Longitude',
            'Rock Specimen Image', 'Outcrop Image',
            'Student ID', 'Submitted By', 'Verified By', 'Status', 'Created At', 'Updated At'
        ])
        
        # Write data rows
        for rock in rocks:
            # Check if images exist and show "(image)" if they do
            rock_specimen_img = '(image)' if rock.get('has_rock_specimen') else ''
            outcrop_img = '(image)' if rock.get('has_outcrop') else ''
            
            writer.writerow([
                rock.get('rock_index', ''),
                rock.get('rock_id', ''),
                rock.get('rock_type', ''),
                rock.get('description', ''),
                rock.get('formation', ''),
                rock.get('location_name', ''),
                rock.get('barangay', ''),
                rock.get('province', ''),
                rock.get('latitude', ''),
                rock.get('longitude', ''),
                rock_specimen_img,
                outcrop_img,
                rock.get('student_id', ''),
                rock.get('student_name', ''),
                rock.get('verified_by_name', ''),
                rock.get('status', ''),
                rock.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('created_at') else '',
                rock.get('updated_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('updated_at') else ''
            ])
        
        # Prepare file for download
        output.seek(0)
        status_suffix = f"_{status_filter}" if status_filter else "_all"
        filename = f"rock_samples{status_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        close_connection(conn)
        
        # Create BytesIO object for CSV
        mem = io.BytesIO()
        mem.write(output.getvalue().encode('utf-8'))
        mem.seek(0)
        
        return send_file(
            mem,
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Error exporting rocks to CSV: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while exporting data.', 'danger')
        return redirect(url_for('admin_rock_list'))

@app.route('/admin/export-rocks/excel')
@login_required
@role_required('admin')
def admin_export_rocks_excel():
    """Export rock samples to Excel format based on current filters and status"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Get filter parameters from request
        search_query = request.args.get('search', '').strip()
        rock_type_filter = request.args.get('rock_type', '').strip()
        status_filter = request.args.get('status', '').strip()  # verified, pending, rejected, or empty for all
        
        # Get filtered rocks with image data for Excel export
        rocks = get_filtered_admin_rocks(conn, search_query, rock_type_filter, status_filter, include_image_data=True)
        
        # Create Excel workbook in memory
        wb = Workbook()
        ws = wb.active
        ws.title = "Rock Samples"
        
        # Define header row (including image columns)
        headers = [
            'Sample Index', 'Rock ID', 'Rock Type', 'Description', 'Formation',
            'Location Name', 'Barangay', 'Province', 'Latitude', 'Longitude',
            'Rock Specimen Image', 'Outcrop Image',
            'Student ID', 'Submitted By', 'Verified By', 'Status', 'Created At', 'Updated At'
        ]
        
        # Style for header row
        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write header row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Set column widths for image columns (columns K and L) - wider for images
        ws.column_dimensions['K'].width = 25  # Rock Specimen Image
        ws.column_dimensions['L'].width = 25  # Outcrop Image
        
        # Set row height for image rows (higher to accommodate images)
        image_row_height = 115  # points
        
        # Write data rows and embed images
        for row_num, rock in enumerate(rocks, 2):
            # Write text data
            ws.cell(row=row_num, column=1, value=rock.get('rock_index', ''))
            ws.cell(row=row_num, column=2, value=rock.get('rock_id', ''))
            ws.cell(row=row_num, column=3, value=rock.get('rock_type', ''))
            ws.cell(row=row_num, column=4, value=rock.get('description', ''))
            ws.cell(row=row_num, column=5, value=rock.get('formation', ''))
            ws.cell(row=row_num, column=6, value=rock.get('location_name', ''))
            ws.cell(row=row_num, column=7, value=rock.get('barangay', ''))
            ws.cell(row=row_num, column=8, value=rock.get('province', ''))
            ws.cell(row=row_num, column=9, value=rock.get('latitude', ''))
            ws.cell(row=row_num, column=10, value=rock.get('longitude', ''))
            
            # Set row height for images
            ws.row_dimensions[row_num].height = image_row_height
            
            # Column 11: Rock Specimen Image
            if rock.get('rock_specimen_data'):
                try:
                    # Create image from bytes
                    img_data = rock.get('rock_specimen_data')
                    # Ensure img_data is bytes
                    if not isinstance(img_data, bytes):
                        if isinstance(img_data, bytearray):
                            img_data = bytes(img_data)
                        else:
                            raise ValueError(f"Image data is not bytes, got {type(img_data)}")
                    img_bytes = io.BytesIO(img_data)
                    
                    # Resize image while maintaining aspect ratio using PIL
                    pil_img = PILImage.open(img_bytes)
                    max_size = 150
                    
                    # Store original dimensions
                    original_width = pil_img.width
                    original_height = pil_img.height
                    
                    # Calculate new size maintaining aspect ratio
                    if original_width > max_size or original_height > max_size:
                        # Use LANCZOS resampling for high quality
                        try:
                            pil_img.thumbnail((max_size, max_size), PILImage.Resampling.LANCZOS)
                        except (AttributeError, TypeError):
                            # Fallback for older PIL versions
                            pil_img.thumbnail((max_size, max_size), PILImage.LANCZOS)
                    
                    # Get final dimensions after thumbnail
                    final_width = pil_img.width
                    final_height = pil_img.height
                    
                    # Save resized image to BytesIO
                    resized_bytes = io.BytesIO()
                    # Determine best format - preserve PNG if possible, otherwise use JPEG
                    if pil_img.mode == 'RGBA':
                        pil_img.save(resized_bytes, format='PNG')
                    elif pil_img.format and pil_img.format.upper() in ['PNG', 'JPEG', 'JPG']:
                        pil_img.save(resized_bytes, format=pil_img.format, quality=85 if pil_img.format.upper() == 'JPEG' else None)
                    else:
                        if pil_img.mode not in ['RGB', 'L']:
                            pil_img = pil_img.convert('RGB')
                        pil_img.save(resized_bytes, format='JPEG', quality=85)
                    resized_bytes.seek(0)
                    
                    # Create openpyxl image
                    img = OpenpyxlImage(resized_bytes)
                    img.width = final_width
                    img.height = final_height
                    
                    # Anchor image to cell
                    img.anchor = f'K{row_num}'
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error adding rock specimen image for sample {rock.get('sample_id')}: {e}")
                    import traceback
                    traceback.print_exc()
                    ws.cell(row=row_num, column=11, value='(image error)')
            else:
                ws.cell(row=row_num, column=11, value='')
            
            # Column 12: Outcrop Image
            if rock.get('outcrop_data'):
                try:
                    # Create image from bytes
                    img_data = rock.get('outcrop_data')
                    # Ensure img_data is bytes
                    if not isinstance(img_data, bytes):
                        if isinstance(img_data, bytearray):
                            img_data = bytes(img_data)
                        else:
                            raise ValueError(f"Image data is not bytes, got {type(img_data)}")
                    img_bytes = io.BytesIO(img_data)
                    
                    # Resize image while maintaining aspect ratio using PIL
                    pil_img = PILImage.open(img_bytes)
                    max_size = 150
                    
                    # Store original dimensions
                    original_width = pil_img.width
                    original_height = pil_img.height
                    
                    # Calculate new size maintaining aspect ratio
                    if original_width > max_size or original_height > max_size:
                        # Use LANCZOS resampling for high quality
                        try:
                            pil_img.thumbnail((max_size, max_size), PILImage.Resampling.LANCZOS)
                        except (AttributeError, TypeError):
                            # Fallback for older PIL versions
                            pil_img.thumbnail((max_size, max_size), PILImage.LANCZOS)
                    
                    # Get final dimensions after thumbnail
                    final_width = pil_img.width
                    final_height = pil_img.height
                    
                    # Save resized image to BytesIO
                    resized_bytes = io.BytesIO()
                    # Determine best format - preserve PNG if possible, otherwise use JPEG
                    if pil_img.mode == 'RGBA':
                        pil_img.save(resized_bytes, format='PNG')
                    elif pil_img.format and pil_img.format.upper() in ['PNG', 'JPEG', 'JPG']:
                        pil_img.save(resized_bytes, format=pil_img.format, quality=85 if pil_img.format.upper() == 'JPEG' else None)
                    else:
                        if pil_img.mode not in ['RGB', 'L']:
                            pil_img = pil_img.convert('RGB')
                        pil_img.save(resized_bytes, format='JPEG', quality=85)
                    resized_bytes.seek(0)
                    
                    # Create openpyxl image
                    img = OpenpyxlImage(resized_bytes)
                    img.width = final_width
                    img.height = final_height
                    
                    # Anchor image to cell
                    img.anchor = f'L{row_num}'
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error adding outcrop image for sample {rock.get('sample_id')}: {e}")
                    import traceback
                    traceback.print_exc()
                    ws.cell(row=row_num, column=12, value='(image error)')
            else:
                ws.cell(row=row_num, column=12, value='')
            
            # Continue with remaining text columns
            ws.cell(row=row_num, column=13, value=rock.get('student_id', ''))
            ws.cell(row=row_num, column=14, value=rock.get('student_name', ''))
            ws.cell(row=row_num, column=15, value=rock.get('verified_by_name', ''))
            ws.cell(row=row_num, column=16, value=rock.get('status', ''))
            ws.cell(row=row_num, column=17, value=rock.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('created_at') else '')
            ws.cell(row=row_num, column=18, value=rock.get('updated_at').strftime('%Y-%m-%d %H:%M:%S') if rock.get('updated_at') else '')
        
        # Auto-adjust column widths for text columns (skip image columns K and L)
        for col_num in range(1, len(headers) + 1):
            if col_num not in [11, 12]:  # Skip image columns
                column_letter = ws.cell(row=1, column=col_num).column_letter
                max_length = 0
                for row in ws.iter_rows(min_row=1, max_row=min(len(rocks) + 1, 100), min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                adjusted_width = min(max_length + 2, 50)
                if adjusted_width > 0:
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        status_suffix = f"_{status_filter}" if status_filter else "_all"
        filename = f"rock_samples{status_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        close_connection(conn)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if conn:
            try:
                close_connection(conn)
            except:
                pass
        print(f"Error exporting rocks to Excel: {e}")
        import traceback
        traceback.print_exc()
        flash('An error occurred while exporting data.', 'danger')
        return redirect(url_for('admin_rock_list'))

@app.route('/admin/add-rock', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_add_rock():
    """Add new rock sample (admin adds their own rock)"""
    if request.method == 'POST':
        user_id = session['user_id']  # Admin's ID
        rock_index = request.form.get('rock_index')
        rock_id = request.form.get('rock_id')
        rock_type = request.form.get('rock_type')
        description = request.form.get('description', '')
        formation = request.form.get('formation', '')
        location_name = request.form.get('location_name')
        barangay = request.form.get('barangay') or None
        province = request.form.get('province') or None
        
        # Handle optional latitude/longitude
        latitude_str = (request.form.get('latitude', '') or '').strip()
        longitude_str = (request.form.get('longitude', '') or '').strip()

        if not latitude_str or not longitude_str:
            flash('Latitude and longitude are required.', 'danger')
            return redirect(url_for('admin_add_rock'))
        try:
            latitude = float(latitude_str)
            longitude = float(longitude_str)
        except ValueError:
            flash('Latitude and longitude must be numeric values.', 'danger')
            return redirect(url_for('admin_add_rock'))

        if latitude < -90 or latitude > 90:
            flash('Latitude must be between -90 and 90 degrees.', 'danger')
            return redirect(url_for('admin_add_rock'))
        if longitude < -180 or longitude > 180:
            flash('Longitude must be between -180 and 180 degrees.', 'danger')
            return redirect(url_for('admin_add_rock'))
        
        rock_specimen = request.files.get('rock_specimen')
        outcrop_image = request.files.get('outcrop_image')
        
        conn = get_db_connection()
        try:
            ensure_rock_location_columns(conn)
            # Insert rock sample - admin rocks are auto-verified
            sample_id = execute_query(conn,
                """INSERT INTO rock_samples (user_id, rock_index, rock_id, rock_type, 
                   description, formation, location_name, barangay, province, latitude, longitude, 
                   status, verified_by, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'verified', %s, NOW())""",
                (user_id, rock_index, rock_id, rock_type, description, formation, 
                 location_name, barangay, province, latitude, longitude, user_id))
            
            # Insert rock specimen image (ensure only one image of this type)
            if rock_specimen and rock_specimen.filename:
                # Delete any existing rock_specimen images for this sample (safety check)
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'rock_specimen'""",
                    (sample_id,))
                rock_specimen_data = rock_specimen.read()
                if rock_specimen_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                           file_size, mime_type, created_at)
                           VALUES (%s, 'rock_specimen', %s, %s, %s, %s, NOW())""",
                        (sample_id, rock_specimen_data, rock_specimen.filename,
                             len(rock_specimen_data), rock_specimen.content_type or 'application/octet-stream'))
            
            # Insert outcrop image (ensure only one image of this type)
            if outcrop_image and outcrop_image.filename:
                # Delete any existing outcrop images for this sample (safety check)
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'outcrop'""",
                    (sample_id,))
                outcrop_image_data = outcrop_image.read()
                if outcrop_image_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name, 
                           file_size, mime_type, created_at)
                           VALUES (%s, 'outcrop', %s, %s, %s, %s, NOW())""",
                        (sample_id, outcrop_image_data, outcrop_image.filename,
                             len(outcrop_image_data), outcrop_image.content_type or 'application/octet-stream'))
            
            # Log activity
            log_activity(conn, session['user_id'], 'submitted', 
                        f'Admin added rock sample: {rock_id} - {rock_type}', sample_id)
            
            close_connection(conn)
            flash('Rock sample added successfully and automatically verified!', 'success')
            return redirect(url_for('admin_rock_list'))
        except Exception as e:
            close_connection(conn)
            flash(f'Error adding rock sample: {str(e)}', 'danger')
    
    return render_template('admin/add_rock.html')

@app.route('/admin/rock-detail/<int:sample_id>')
@login_required
@role_required('admin')
def admin_rock_detail(sample_id):
    """View detailed information about a rock sample"""
    conn = get_db_connection()
    
    rock = fetch_one(conn,
        """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as student_name,
           u.email as student_email, u.school_id,
           CONCAT(v.first_name, ' ', v.last_name) as verified_by_name
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           LEFT JOIN users v ON rs.verified_by = v.user_id
           WHERE rs.sample_id = %s""",
        (sample_id,))
    
    images = fetch_all(conn,
        """SELECT image_id, image_type, file_name, file_size, mime_type, created_at
           FROM images WHERE sample_id = %s""",
        (sample_id,))
    
    approval_history = fetch_all(conn,
        """SELECT al.*, CONCAT(u.first_name, ' ', u.last_name) as user_name
           FROM approval_logs al
           JOIN users u ON al.user_id = u.user_id
           WHERE al.sample_id = %s
           ORDER BY al.timestamp DESC""",
        (sample_id,))
    
    close_connection(conn)
    return render_template('admin/rock_detail_view.html', rock=rock, 
                         images=images, approval_history=approval_history)

@app.route('/admin/edit-rock/<int:sample_id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_edit_rock(sample_id):
    """Edit a rock sample"""
    conn = get_db_connection()
    ensure_rock_location_columns(conn)
    
    if request.method == 'POST':
        try:
            # Get form data
            rock_id = request.form.get('rock_id', '').strip()
            rock_type = request.form.get('rock_type', '').strip()
            description = request.form.get('description', '').strip()
            location_name = request.form.get('location_name', '').strip()
            barangay = request.form.get('barangay', '').strip() or None
            province = request.form.get('province', '').strip() or None
            latitude_str = (request.form.get('latitude', '') or '').strip()
            longitude_str = (request.form.get('longitude', '') or '').strip()
            formation = request.form.get('formation', '').strip()
            rock_index = request.form.get('rock_index', '').strip()

            if not latitude_str or not longitude_str:
                close_connection(conn)
                flash('Latitude and longitude are required.', 'error')
                return redirect(url_for('admin_edit_rock', sample_id=sample_id))
            try:
                latitude = float(latitude_str)
                longitude = float(longitude_str)
            except ValueError:
                close_connection(conn)
                flash('Latitude and longitude must be numeric values.', 'error')
                return redirect(url_for('admin_edit_rock', sample_id=sample_id))

            if latitude < -90 or latitude > 90:
                close_connection(conn)
                flash('Latitude must be between -90 and 90 degrees.', 'error')
                return redirect(url_for('admin_edit_rock', sample_id=sample_id))
            if longitude < -180 or longitude > 180:
                close_connection(conn)
                flash('Longitude must be between -180 and 180 degrees.', 'error')
                return redirect(url_for('admin_edit_rock', sample_id=sample_id))
            
            # Validate required fields
            if not rock_id or not rock_type or not location_name:
                flash('Rock ID, Rock Type, and Location are required!', 'error')
                return redirect(url_for('admin_edit_rock', sample_id=sample_id))
            
            # Update the rock sample
            execute_query(conn,
                """UPDATE rock_samples SET 
                   rock_id = %s, rock_type = %s, description = %s,
                   location_name = %s, barangay = %s, province = %s, latitude = %s, longitude = %s,
                   formation = %s, rock_index = %s,
                   updated_at = NOW()
                   WHERE sample_id = %s""",
                (rock_id, rock_type, description, location_name, barangay, province, latitude, longitude,
                 formation, rock_index, sample_id))
            
            # Handle image removals
            remove_ids = request.form.getlist('remove_images')
            if remove_ids:
                for image_id in remove_ids:
                    if image_id.isdigit():
                        execute_query(conn,
                            "DELETE FROM images WHERE image_id = %s AND sample_id = %s",
                            (int(image_id), sample_id))

            # Handle new uploads - delete existing images of same type first to ensure only one
            rock_specimen_file = request.files.get('rock_specimen')
            if rock_specimen_file and rock_specimen_file.filename:
                # Delete existing rock_specimen image before inserting new one
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'rock_specimen'""",
                    (sample_id,))
                specimen_data = rock_specimen_file.read()
                if specimen_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name,
                           file_size, mime_type, created_at)
                           VALUES (%s, 'rock_specimen', %s, %s, %s, %s, NOW())""",
                        (sample_id, specimen_data, rock_specimen_file.filename,
                         len(specimen_data), rock_specimen_file.content_type or 'application/octet-stream'))

            outcrop_file = request.files.get('outcrop_image')
            if outcrop_file and outcrop_file.filename:
                # Delete existing outcrop image before inserting new one
                execute_query(conn,
                    """DELETE FROM images 
                       WHERE sample_id = %s AND image_type = 'outcrop'""",
                    (sample_id,))
                outcrop_data = outcrop_file.read()
                if outcrop_data:
                    execute_query(conn,
                        """INSERT INTO images (sample_id, image_type, image_data, file_name,
                           file_size, mime_type, created_at)
                           VALUES (%s, 'outcrop', %s, %s, %s, %s, NOW())""",
                        (sample_id, outcrop_data, outcrop_file.filename,
                         len(outcrop_data), outcrop_file.content_type or 'application/octet-stream'))
            
            # Log the edit activity
            user_id = session['user_id']
            execute_query(conn,
                """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
                   VALUES (%s, %s, 'edited', %s)""",
                (user_id, sample_id, f'Rock sample edited by admin: {rock_id}'))
            
            close_connection(conn)
            flash('Rock sample updated successfully!', 'success')
            return redirect(url_for('admin_rock_detail', sample_id=sample_id))
            
        except Exception as e:
            print(f"Error updating rock sample {sample_id}: {str(e)}")
            flash('Error updating rock sample!', 'error')
            close_connection(conn)
            return redirect(url_for('admin_edit_rock', sample_id=sample_id))
    
    else:
        # GET request - show edit form
        rock = fetch_one(conn,
            """SELECT rs.*, CONCAT(u.first_name, ' ', u.last_name) as student_name,
               u.email as student_email, u.school_id
               FROM rock_samples rs
               JOIN users u ON rs.user_id = u.user_id
               WHERE rs.sample_id = %s""",
            (sample_id,))
        
        if not rock:
            close_connection(conn)
            flash('Rock sample not found!', 'error')
            return redirect(url_for('admin_rock_list'))
        
        # Fetch associated images for preview in edit screen
        images = fetch_all(conn,
            """SELECT image_id, image_type, file_name, file_size, mime_type, created_at
               FROM images
               WHERE sample_id = %s
               ORDER BY created_at DESC""",
            (sample_id,))
        
        close_connection(conn)
        return render_template('admin/edit_rock.html', rock=rock, images=images)

@app.route('/admin/archive-rock/<int:sample_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_archive_rock(sample_id):
    """Archive a rock sample"""
    conn = None
    try:
        reason = request.form.get('reason', 'Archived by admin')
        user_id = session['user_id']
        
        conn = get_db_connection()
        
        # Check if rock sample exists
        rock = fetch_one(conn, "SELECT * FROM rock_samples WHERE sample_id = %s", (sample_id,))
        if not rock:
            close_connection(conn)
            flash('Rock sample not found!', 'error')
            return redirect(url_for('admin_rock_list'))
        
        # Check if already archived
        existing_archive = fetch_one(conn, "SELECT * FROM archives WHERE sample_id = %s", (sample_id,))
        if existing_archive:
            close_connection(conn)
            flash('Rock sample is already archived!', 'warning')
            return redirect(url_for('admin_rock_list'))
        
        # Archive the rock sample
        archive_id = execute_query(conn,
            """INSERT INTO archives (sample_id, archived_by, archive_reason, status)
               VALUES (%s, %s, %s, 'archived')""",
            (sample_id, user_id, reason))
        
        # Log the activity
        activity_id = execute_query(conn,
            """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
               VALUES (%s, %s, 'archived', %s)""",
            (user_id, sample_id, f'Rock sample archived by admin: {reason}'))
        
        close_connection(conn)
        flash('Rock sample archived successfully!', 'success')
        
    except Exception as e:
        print(f"Error archiving rock sample {sample_id}: {str(e)}")
        flash('Error archiving rock sample!', 'error')
        # Ensure connection is closed even on error
        if conn:
            try:
                close_connection(conn)
            except:
                pass
    
    return redirect(url_for('admin_rock_list'))

@app.route('/admin/archives')
@login_required
@role_required('admin')
def admin_archives():
    """View all archived rock samples"""
    conn = get_db_connection()
    
    archives = fetch_all(conn,
        """SELECT a.*, rs.rock_id, rs.rock_type, rs.location_name,
           CONCAT(u.first_name, ' ', u.last_name) as archived_by_name,
           CONCAT(s.first_name, ' ', s.last_name) as student_name
           FROM archives a
           JOIN rock_samples rs ON a.sample_id = rs.sample_id
           JOIN users u ON a.archived_by = u.user_id
           JOIN users s ON rs.user_id = s.user_id
           ORDER BY a.archived_at DESC""")
    
    close_connection(conn)
    return render_template('admin/archives.html', archives=archives)

@app.route('/admin/unarchive/<int:sample_id>', methods=['POST'])
@login_required
@role_required('admin')
def admin_unarchive_rock(sample_id):
    """Unarchive a rock sample"""
    conn = None
    try:
        user_id = session['user_id']
        
        conn = get_db_connection()
        
        # Check if rock sample exists
        rock = fetch_one(conn, "SELECT * FROM rock_samples WHERE sample_id = %s", (sample_id,))
        if not rock:
            close_connection(conn)
            flash('Rock sample not found!', 'error')
            return redirect(url_for('admin_archives'))
        
        # Check if archived
        existing_archive = fetch_one(conn, "SELECT * FROM archives WHERE sample_id = %s", (sample_id,))
        if not existing_archive:
            close_connection(conn)
            flash('Rock sample is not archived!', 'warning')
            return redirect(url_for('admin_archives'))
        
        # Delete the archive record
        execute_query(conn,
            "DELETE FROM archives WHERE sample_id = %s",
            (sample_id,))
        
        # Log the activity
        execute_query(conn,
            """INSERT INTO activity_logs (user_id, sample_id, activity_type, description)
               VALUES (%s, %s, 'edited', %s)""",
            (user_id, sample_id, f'Rock sample unarchived by admin'))
        
        close_connection(conn)
        flash('Rock sample unarchived successfully!', 'success')
        
    except Exception as e:
        print(f"Error unarchiving rock sample {sample_id}: {str(e)}")
        flash('Error unarchiving rock sample!', 'error')
        # Ensure connection is closed even on error
        if conn:
            try:
                close_connection(conn)
            except:
                pass
    
    return redirect(url_for('admin_archives'))

@app.route('/admin/map')
@login_required
@role_required('admin')
def admin_map():
    """Interactive map showing all rock sample locations"""
    conn = get_db_connection()
    
    # Get all rock samples with coordinates
    rocks = fetch_all(conn,
        """SELECT rs.sample_id, rs.rock_id, rs.rock_type, rs.formation, rs.location_name, 
           rs.latitude, rs.longitude, rs.status, rs.created_at,
           CONCAT(u.first_name, ' ', u.last_name) as student_name
           FROM rock_samples rs
           JOIN users u ON rs.user_id = u.user_id
           WHERE rs.latitude IS NOT NULL AND rs.longitude IS NOT NULL
           ORDER BY rs.created_at DESC""")
    
    # Get city-level statistics
    cities = fetch_all(conn,
        """SELECT location_name, COUNT(*) as specimen_count,
           AVG(latitude) as avg_lat, AVG(longitude) as avg_lng
           FROM rock_samples 
           WHERE latitude IS NOT NULL AND longitude IS NOT NULL
           GROUP BY location_name""")
    
    close_connection(conn)
    return render_template('admin/map.html', rocks=rocks, cities=cities)

@app.route('/admin/activity-logs')
@login_required
@role_required('admin')
def admin_activity_logs():
    """View all activity logs with filtering"""
    conn = get_db_connection()
    
    # Get filter parameters
    user_filter = request.args.get('user', '').strip()
    action_filter = request.args.get('action', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to = request.args.get('date_to', '').strip()
    
    # Build the base query
    base_query = """SELECT al.*, CONCAT(u.first_name, ' ', u.last_name) as user_name,
           u.username, u.role as user_role, rs.rock_id, rs.rock_type
           FROM activity_logs al
           JOIN users u ON al.user_id = u.user_id
           LEFT JOIN rock_samples rs ON al.sample_id = rs.sample_id"""
    
    # Build WHERE conditions
    where_conditions = []
    params = []
    
    # Add user filter
    if user_filter:
        where_conditions.append("al.user_id = %s")
        params.append(user_filter)
    
    # Add action filter
    if action_filter:
        where_conditions.append("al.activity_type = %s")
        params.append(action_filter)
    
    # Add date filters
    if date_from:
        where_conditions.append("DATE(al.timestamp) >= %s")
        params.append(date_from)
    
    if date_to:
        where_conditions.append("DATE(al.timestamp) <= %s")
        params.append(date_to)
    
    # Combine query
    if where_conditions:
        query = base_query + " WHERE " + " AND ".join(where_conditions)
    else:
        query = base_query
    
    query += " ORDER BY al.timestamp DESC LIMIT 200"
    
    logs = fetch_all(conn, query, params)
    
    # Get users for filter dropdown
    users = fetch_all(conn, "SELECT user_id, username FROM users WHERE is_active = TRUE ORDER BY username")
    
    close_connection(conn)
    return render_template('admin/activity_logs.html', logs=logs, users=users,
                          user_filter=user_filter, action_filter=action_filter,
                          date_from=date_from, date_to=date_to)

@app.route('/admin/settings', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_settings():
    """Admin settings page"""
    conn = get_db_connection()
    user_id = session['user_id']
    
    ensure_user_photo_columns(conn)
    
    if request.method == 'POST':
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        email = request.form.get('email')
        new_password = request.form.get('new_password')
        
        if new_password:
            password_hash = generate_password_hash(new_password)
            execute_query(conn,
                """UPDATE users SET first_name = %s, last_name = %s, email = %s, 
                   password_hash = %s WHERE user_id = %s""",
                (first_name, last_name, email, password_hash, user_id))
        else:
            execute_query(conn,
                """UPDATE users SET first_name = %s, last_name = %s, email = %s 
                   WHERE user_id = %s""",
                (first_name, last_name, email, user_id))
        
        flash('Settings updated successfully!', 'success')
        return redirect(url_for('admin_settings'))
    
    user = fetch_one(conn, "SELECT * FROM users WHERE user_id = %s", (user_id,))
    close_connection(conn)
    return render_template('admin/settings.html', user=user)

@app.route('/admin/upload-photo', methods=['POST'])
@login_required
@role_required('admin')
def admin_upload_photo():
    """Upload or replace admin profile photo"""
    conn = get_db_connection()
    try:
        ensure_user_photo_columns(conn)
        file = request.files.get('profile_photo')
        if not file or not file.filename:
            close_connection(conn)
            flash('Please select an image to upload', 'danger')
            return redirect(url_for('admin_settings'))
        filename = secure_filename(file.filename)
        data = file.read()
        if not data:
            close_connection(conn)
            flash('Empty file uploaded', 'danger')
            return redirect(url_for('admin_settings'))
        execute_query(conn, """
            UPDATE users
            SET profile_image = %s,
                profile_image_mime = %s,
                profile_image_name = %s,
                profile_image_size = %s,
                updated_at = NOW()
            WHERE user_id = %s
        """, (data, file.content_type or 'application/octet-stream', filename, len(data), session['user_id']))
        log_activity(conn, session['user_id'], 'profile_photo_updated', 'Updated profile photo')
        close_connection(conn)
        flash('Profile photo updated', 'success')
    except Exception as e:
        try:
            close_connection(conn)
        except Exception:
            pass
        flash(f'Error uploading photo: {str(e)}', 'danger')
    return redirect(url_for('admin_settings'))

# ============================================================================
# IMAGE ROUTES
# ============================================================================

@app.route('/image/<int:image_id>')
@login_required
def serve_image(image_id):
    """Serve an image from the database"""
    conn = get_db_connection()
    
    image = fetch_one(conn,
        "SELECT image_data, mime_type, file_name FROM images WHERE image_id = %s",
        (image_id,))
    
    close_connection(conn)
    
    if image:
        return send_file(
            io.BytesIO(image['image_data']),
            mimetype=image['mime_type'],
            as_attachment=False,
            download_name=image['file_name']
        )
    else:
        flash('Image not found', 'danger')
        return redirect(url_for('index'))

@app.route('/image/sample/<int:sample_id>/<image_type>')
@login_required
def serve_sample_image(sample_id, image_type):
    """Serve a specific type of image for a rock sample"""
    conn = get_db_connection()
    
    image = fetch_one(conn,
        """SELECT image_data, mime_type, file_name 
           FROM images 
           WHERE sample_id = %s AND image_type = %s
           ORDER BY created_at DESC LIMIT 1""",
        (sample_id, image_type))
    
    close_connection(conn)
    
    if image:
        return send_file(
            io.BytesIO(image['image_data']),
            mimetype=image['mime_type'],
            as_attachment=False,
            download_name=image['file_name']
        )
    else:
        # Return placeholder image
        return redirect(url_for('static', filename='images/no-image.png'))

@app.route('/user/photo/<int:user_id>')
@login_required
def serve_user_photo(user_id):
    """Serve a user's profile photo; falls back to placeholder if none."""
    conn = get_db_connection()
    ensure_user_photo_columns(conn)
    image = fetch_one(conn, """
        SELECT profile_image, profile_image_mime, profile_image_name
        FROM users
        WHERE user_id = %s
    """, (user_id,))
    close_connection(conn)
    if image and image.get('profile_image'):
        return send_file(
            io.BytesIO(image['profile_image']),
            mimetype=image.get('profile_image_mime') or 'image/jpeg',
            as_attachment=False,
            download_name=image.get('profile_image_name') or 'profile.jpg'
        )
    # No profile image stored; return 404 so clients can show local fallback (white circle)
    abort(404)

# ============================================================================
# ERROR HANDLERS
# ============================================================================

@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    return render_template('500.html'), 500

@app.errorhandler(403)
def forbidden(error):
    """Handle 403 errors"""
    flash('You do not have permission to access this page', 'danger')
    return redirect(url_for('index'))

# ============================================================================
# ADDITIONAL ADMIN ROUTES
# ============================================================================

@app.route('/admin/add-user', methods=['POST'])
@login_required
@role_required('admin')
def admin_add_user():
    """Add a new user"""
    try:
        username = request.form.get('username')
        email = request.form.get('email')
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        role = request.form.get('role', 'student')
        school_id = request.form.get('school_id')
        password = request.form.get('password')
        
        if not all([username, email, first_name, last_name, password]):
            flash('All fields are required', 'error')
            return redirect(url_for('admin_manage_users'))
        
        conn = get_db_connection()
        
        # Check if username or email already exists
        existing_user = fetch_one(conn,
            "SELECT user_id FROM users WHERE username = %s OR email = %s",
            (username, email))
        
        if existing_user:
            flash('Username or email already exists', 'error')
            close_connection(conn)
            return redirect(url_for('admin_manage_users'))
        
        # Hash password and insert user
        password_hash = generate_password_hash(password)
        user_id = execute_query(conn,
            """INSERT INTO users (username, email, password_hash, first_name, last_name, 
               role, school_id, is_active, created_at) 
               VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE, NOW())""",
            (username, email, password_hash, first_name, last_name, role, school_id))
        
        log_activity(conn, session['user_id'], 'user_created', 
                    f'Created user: {username} ({role})')
        
        close_connection(conn)
        flash(f'User {username} created successfully', 'success')
        
    except Exception as e:
        flash(f'Error creating user: {str(e)}', 'error')
        if 'conn' in locals():
            close_connection(conn)
    
    return redirect(url_for('admin_manage_users'))

@app.route('/admin/edit-user/<int:user_id>', methods=['GET', 'POST'])
@login_required
@role_required('admin')
def admin_edit_user(user_id):
    """Edit user information"""
    conn = get_db_connection()
    
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            email = request.form.get('email')
            full_name = request.form.get('full_name', '')
            # Derive first_name and last_name from full_name for backward compatibility
            full_name = (full_name or '').strip()
            if ' ' in full_name:
                first_name, last_name = full_name.split(' ', 1)
            else:
                first_name, last_name = full_name, ''
            role = request.form.get('role')
            school_id = request.form.get('school_id')
            is_active = request.form.get('is_active') == 'on'
            
            if not all([username, email, first_name, role]):
                flash('All required fields must be filled', 'error')
                close_connection(conn)
                return redirect(url_for('admin_edit_user', user_id=user_id))
            
            # Check if username or email already exists for other users
            existing_user = fetch_one(conn,
                "SELECT user_id FROM users WHERE (username = %s OR email = %s) AND user_id != %s",
                (username, email, user_id))
            
            if existing_user:
                flash('Username or email already exists for another user', 'error')
                close_connection(conn)
                return redirect(url_for('admin_edit_user', user_id=user_id))
            
            # Update user information
            execute_query(conn,
                """UPDATE users SET username = %s, email = %s, first_name = %s, 
                   last_name = %s, role = %s, school_id = %s, is_active = %s, updated_at = NOW()
                   WHERE user_id = %s""",
                (username, email, first_name, last_name, role, school_id, is_active, user_id))
            
            # Update password if provided
            new_password = request.form.get('new_password')
            if new_password and new_password.strip():
                password_hash = generate_password_hash(new_password)
                execute_query(conn,
                    "UPDATE users SET password_hash = %s WHERE user_id = %s",
                    (password_hash, user_id))
            
            log_activity(conn, session['user_id'], 'user_updated', 
                        f'Updated user: {username} ({role})')
            
            close_connection(conn)
            flash(f'User {username} updated successfully', 'success')
            return redirect(url_for('admin_manage_users'))
            
        except Exception as e:
            flash(f'Error updating user: {str(e)}', 'error')
            if 'conn' in locals():
                close_connection(conn)
            return redirect(url_for('admin_edit_user', user_id=user_id))
    
    else:  # GET request
        try:
            # Get user information
            user = fetch_one(conn,
                "SELECT * FROM users WHERE user_id = %s",
                (user_id,))
            
            if not user:
                flash('User not found', 'error')
                close_connection(conn)
                return redirect(url_for('admin_manage_users'))
            
            close_connection(conn)
            return render_template('admin/edit_user.html', user=user)
            
        except Exception as e:
            flash(f'Error loading user: {str(e)}', 'error')
            if 'conn' in locals():
                close_connection(conn)
            return redirect(url_for('admin_manage_users'))

# Removed export-to-CSV endpoint for admin

@app.route('/admin/update-profile', methods=['POST'])
@login_required
@role_required('admin')
def admin_update_profile():
    """Update admin profile"""
    try:
        first_name = request.form.get('first_name')
        last_name = request.form.get('last_name')
        email = request.form.get('email')
        
        conn = get_db_connection()
        
        # Check if email is already taken by another user
        existing_user = fetch_one(conn,
            "SELECT user_id FROM users WHERE email = %s AND user_id != %s",
            (email, session['user_id']))
        
        if existing_user:
            flash('Email already taken by another user', 'error')
            close_connection(conn)
            return redirect(url_for('admin_settings'))
        
        # Update user profile
        execute_query(conn,
            "UPDATE users SET first_name = %s, last_name = %s, email = %s, updated_at = NOW() WHERE user_id = %s",
            (first_name, last_name, email, session['user_id']))
        
        # Update session
        session['full_name'] = f"{first_name} {last_name}"
        
        log_activity(conn, session['user_id'], 'profile_updated', 'Updated profile information')
        
        close_connection(conn)
        flash('Profile updated successfully', 'success')
        
    except Exception as e:
        flash(f'Error updating profile: {str(e)}', 'error')
        if 'conn' in locals():
            close_connection(conn)
    
    return redirect(url_for('admin_settings'))

@app.route('/admin/change-password', methods=['POST'])
@login_required
@role_required('admin')
def admin_change_password():
    """Change admin password"""
    try:
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if new_password != confirm_password:
            flash('New passwords do not match', 'error')
            return redirect(url_for('admin_settings'))
        
        conn = get_db_connection()
        
        # Verify current password
        user = fetch_one(conn, "SELECT password_hash FROM users WHERE user_id = %s", (session['user_id'],))
        
        if not user or not check_password_hash(user['password_hash'], current_password):
            flash('Current password is incorrect', 'error')
            close_connection(conn)
            return redirect(url_for('admin_settings'))
        
        # Update password
        new_hash = generate_password_hash(new_password)
        execute_query(conn,
            "UPDATE users SET password_hash = %s, updated_at = NOW() WHERE user_id = %s",
            (new_hash, session['user_id']))
        
        log_activity(conn, session['user_id'], 'password_changed', 'Changed password')
        
        close_connection(conn)
        flash('Password changed successfully', 'success')
        
    except Exception as e:
        flash(f'Error changing password: {str(e)}', 'error')
        if 'conn' in locals():
            close_connection(conn)
    
    return redirect(url_for('admin_settings'))


# ============================================================================
# ERROR HANDLERS
# ============================================================================

# ============================================================================
# RUN APPLICATION
# ============================================================================

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5817)

